#!/usr/bin/env python3
"""Generate and execute an AppleScript to populate c1v_QFD.xlsx from c1v_QFD.json.

Preserves template formatting (AppleScript path; openpyxl would destroy merged cells / borders).
"""
import json
import subprocess
from pathlib import Path
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
XLSX = HERE / "c1v_QFD.xlsx"
JSON_PATH = HERE / "c1v_QFD.json"

with open(JSON_PATH) as f:
    data = json.load(f)

writes: list[tuple[str, object]] = []

# --- Metadata ---
md = data["metadata"]
writes += [
    ("B2", md["project_title"]),
    ("B4", md["developed_by"]),
    ("B6", md["last_updated"]),
]

# --- Front Porch (PCs in rows 33-38; B=full_attribute, C=short_name, D=weight) ---
for i, pc in enumerate(data["front_porch"]):
    row = 33 + i
    writes += [
        (f"B{row}", pc["full_attribute"]),
        (f"C{row}", pc["short_name"]),
        (f"D{row}", pc["relative_importance"]),
    ]

# --- Second Floor (EC name row 30, direction row 31, column n+4 = E for EC1..V for EC18) ---
def ec_col(n: int) -> str:
    return get_column_letter(n + 4)

for ec in data["second_floor"]:
    n = ec["ec_id"]
    col = ec_col(n)
    writes += [
        (f"{col}30", ec["name"]),
        (f"{col}31", ec["direction_of_change"]),
    ]

# --- Main Floor (PC rows 33-38 x EC cols E..V) ---
pc_order = ["PC.1", "PC.2", "PC.3", "PC.4", "PC.5", "PC.6"]
for pc_i, pc_key in enumerate(pc_order):
    row = 33 + pc_i
    for n in range(1, 19):
        val = data["main_floor"][pc_key][f"EC{n}"]
        if val != 0:
            writes.append((f"{ec_col(n)}{row}", val))

# --- Roof (lower triangle): EC_i x EC_j with i<j -> row=j+2, col=i+4 ---
for key, v in data["roof"].items():
    if key.startswith("_"):
        continue
    i, j = map(int, key.replace("EC", "").split("_"))
    assert i < j
    cell = f"{get_column_letter(i + 4)}{j + 2}"
    writes.append((cell, v))

# --- Back Porch (AE=A_low, AF=A_high, AG=A_target, AH=B, AI=C; rows 33-38) ---
for pc_i, pc_key in enumerate(pc_order):
    row = 33 + pc_i
    bp = data["back_porch"][pc_key]
    writes += [
        (f"AE{row}", bp["A_low"]),
        (f"AF{row}", bp["A_high"]),
        (f"AG{row}", bp["A_target"]),
        (f"AH{row}", bp["B"]),
        (f"AI{row}", bp["C"]),
    ]

# --- Basement ---
writes.append(("C46", data["back_porch"]["_competitor_b_name"]))
writes.append(("C47", data["back_porch"]["_competitor_c_name"]))

for n in range(1, 19):
    col = ec_col(n)
    b = data["basement"][f"EC{n}"]
    writes += [
        (f"{col}45", b["unit"]),
        (f"{col}46", b["competitor_b"]),
        (f"{col}47", b["competitor_c"]),
        (f"{col}48", b["external_threshold"]),
        (f"{col}49", b["target"]),
        (f"{col}53", b["technical_difficulty"]),
        (f"{col}54", b["estimated_cost"]),
    ]

# --- Build AppleScript ---
# AppleScript path requires macOS-style HFS-ish; use POSIX path via the modern API.
xlsx_posix = str(XLSX.resolve())
# Convert POSIX path to HFS path via AppleScript inline.
applescript_lines = [
    'tell application "Microsoft Excel"',
    '    activate',
    f'    set wb to open workbook workbook file name (POSIX file "{xlsx_posix}" as string)',
    '    delay 2',
    '    set ws to sheet "QFD Template" of wb',
]
for cell, val in writes:
    if isinstance(val, str):
        escaped = val.replace('\\', '\\\\').replace('"', '\\"')
        applescript_lines.append(f'    set value of cell "{cell}" of ws to "{escaped}"')
    elif isinstance(val, bool):
        applescript_lines.append(f'    set value of cell "{cell}" of ws to {str(val).lower()}')
    elif isinstance(val, (int, float)):
        applescript_lines.append(f'    set value of cell "{cell}" of ws to {val}')
    else:
        applescript_lines.append(f'    set value of cell "{cell}" of ws to "{val}"')

applescript_lines += [
    '    save wb',
    '    close wb saving no',
    '    return "Wrote ' + str(len(writes)) + ' cells."',
    'end tell',
]

script = "\n".join(applescript_lines)
(HERE / "write_xlsx.applescript").write_text(script)
print(f"AppleScript generated: {len(writes)} writes. Executing...")

result = subprocess.run(
    ["osascript", "-e", script],
    capture_output=True, text=True, timeout=600,
)
print("STDOUT:", result.stdout.strip())
if result.returncode != 0:
    print("STDERR:", result.stderr.strip())
    raise SystemExit(result.returncode)
