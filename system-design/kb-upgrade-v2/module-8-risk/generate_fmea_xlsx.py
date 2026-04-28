#!/usr/bin/env python3
"""Generate fmea_table.xlsx from fmea_table.json + rating_scales.json.

Mirrors the convention set by M5/M6 generators: open the JSON artifacts,
write a multi-sheet xlsx, apply conditional formatting for RPN cells.

Usage: python3 generate_fmea_xlsx.py
"""
from __future__ import annotations

import json
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("openpyxl required: pip install openpyxl")

HERE = Path(__file__).parent
TABLE = json.loads((HERE / "fmea_table.json").read_text())
SCALES = json.loads((HERE / "rating_scales.json").read_text())
STOPLIGHTS = json.loads((HERE / "stoplight_charts.json").read_text())

OUT = HERE / "fmea_table.xlsx"

COL_ORDER = TABLE["_column_order"]

CRIT_COLOR = {
    "HIGH":        "DC2626",
    "MEDIUM HIGH": "EA580C",
    "MEDIUM":      "EAB308",
    "MEDIUM LOW":  "84CC16",
    "LOW":         "16A34A",
}

HEADER_FILL = PatternFill("solid", fgColor="0B2C29")
HEADER_FONT = Font(name="Space Grotesk", color="FBFCFC", bold=True, size=11)
BODY_FONT = Font(name="Consolas", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_master_table(wb: openpyxl.Workbook) -> None:
    ws = wb.active
    ws.title = "FMEA Master"
    headers = [c.replace("_", " ").upper() for c in COL_ORDER]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for row in TABLE["rows"]:
        out = []
        for col in COL_ORDER:
            val = row.get(col, "")
            if isinstance(val, list):
                val = "; ".join(str(x) for x in val)
            out.append(val)
        ws.append(out)
        r = ws.max_row
        for c in range(1, len(COL_ORDER) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
        crit_col = COL_ORDER.index("risk_criticality") + 1
        adj_crit_col = COL_ORDER.index("adjusted_criticality") + 1
        for col_idx in (crit_col, adj_crit_col):
            cell = ws.cell(row=r, column=col_idx)
            color = CRIT_COLOR.get(cell.value, "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(name="Space Grotesk", color="FFFFFF", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "failure_mode_id": 10, "subsystem": 22, "failure_mode": 38, "failure_effects": 40,
        "possible_cause": 42, "mmmme": 10, "interface_ref": 10,
        "severity": 8, "likelihood": 10, "rpn": 6, "risk_criticality": 14,
        "corrective_action": 50,
        "adjusted_severity": 10, "adjusted_likelihood": 11, "adjusted_rpn": 9, "adjusted_criticality": 14,
        "corrective_action_effort": 14,
        "detectability": 11, "troubleshooting": 50,
    }
    for i, col in enumerate(COL_ORDER, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 14)
    ws.freeze_panes = "E2"


def write_scales(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Rating Scales")
    ws.append(["Severity Scale (1-4)"])
    ws.append(["Rating", "Label", "Conditions"])
    for lvl in SCALES["severity"]["levels"]:
        ws.append([lvl["rating"], lvl["label"], "\n".join(lvl["conditions"])])
    ws.append([])
    ws.append(["Likelihood Scale (1-5)"])
    ws.append(["Rating", "Label", "Conditions"])
    for lvl in SCALES["likelihood"]["levels"]:
        ws.append([lvl["rating"], lvl["label"], "\n".join(lvl["conditions"])])
    ws.append([])
    ws.append(["Criticality Ranges (RPN)"])
    ws.append(["RPN Range", "Category", "Color"])
    for r in SCALES["criticality_ranges"]:
        ws.append([f"{r['rpn_min']}-{r['rpn_max']}", r["category"], r["color_name"]])
    for col, width in (("A", 14), ("B", 18), ("C", 60)):
        ws.column_dimensions[col].width = width


def write_stoplight(wb: openpyxl.Workbook, sheet: str, chart: dict) -> None:
    ws = wb.create_sheet(sheet)
    ws.append([chart["description"]])
    ws.append([])
    ws.append(["", "S=1 Negligible", "S=2 Marginal", "S=3 Critical", "S=4 Catastrophic"])
    for L in ("L5", "L4", "L3", "L2", "L1"):
        row = [L]
        for S in ("S1", "S2", "S3", "S4"):
            row.append(chart["matrix"][L][S])
        ws.append(row)
    ws.append([])
    ws.append(["Criticality totals"])
    for cat, entry in chart["criticality_totals"].items():
        ws.append([cat, entry["rpn_range"], entry["count"]])
    for col, width in (("A", 18), ("B", 18), ("C", 18), ("D", 18), ("E", 18)):
        ws.column_dimensions[col].width = width


def main() -> None:
    wb = openpyxl.Workbook()
    write_master_table(wb)
    write_scales(wb)
    write_stoplight(wb, "Stoplight Before", STOPLIGHTS["before_corrective_actions"])
    write_stoplight(wb, "Stoplight After", STOPLIGHTS["after_corrective_actions"])
    wb.save(OUT)
    print(f"Wrote {OUT} ({OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
