#!/usr/bin/env python3
"""Convert PDFs to markdown and Excel files to JSON."""

import os
import json
import pdfplumber
import openpyxl
from pathlib import Path

BASE_DIR = Path("/Users/davidancor/Documents/MDR/c1v/apps/product-helper/prd-doc-product-helper")

def convert_pdf_to_markdown(pdf_path: Path) -> str:
    """Extract text from PDF and format as markdown."""
    text_parts = []

    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            page_text = page.extract_text()
            if page_text:
                text_parts.append(f"## Page {i}\n\n{page_text}")

            # Also extract tables if present
            tables = page.extract_tables()
            for j, table in enumerate(tables, 1):
                if table:
                    text_parts.append(f"\n### Table {j} (Page {i})\n")
                    # Convert table to markdown
                    if table and len(table) > 0:
                        # Header row
                        header = table[0]
                        if header:
                            text_parts.append("| " + " | ".join(str(cell) if cell else "" for cell in header) + " |")
                            text_parts.append("| " + " | ".join("---" for _ in header) + " |")
                        # Data rows
                        for row in table[1:]:
                            if row:
                                text_parts.append("| " + " | ".join(str(cell) if cell else "" for cell in row) + " |")

    return "\n\n".join(text_parts)


def convert_excel_to_json(excel_path: Path) -> dict:
    """Convert Excel file to JSON with all sheets."""
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    result = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = []
        headers = None

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue

            if i == 0:
                # Use first row as headers
                headers = [str(cell) if cell else f"col_{j}" for j, cell in enumerate(row)]
            else:
                if headers:
                    row_dict = {}
                    for j, cell in enumerate(row):
                        if j < len(headers):
                            row_dict[headers[j]] = cell
                    rows.append(row_dict)

        result[sheet_name] = rows

    return result


def find_and_convert_pdfs():
    """Find all PDFs and convert them to markdown."""
    pdfs = list(BASE_DIR.rglob("*.pdf"))

    # Create output directory
    output_dir = BASE_DIR / "converted" / "markdown"
    output_dir.mkdir(parents=True, exist_ok=True)

    for pdf_path in pdfs:
        # Skip zip files that contain pdf in the name
        if ".zip" in str(pdf_path):
            continue

        print(f"Converting PDF: {pdf_path.name}")
        try:
            markdown_content = convert_pdf_to_markdown(pdf_path)

            # Create output filename
            output_name = pdf_path.stem + ".md"
            output_path = output_dir / output_name

            with open(output_path, "w", encoding="utf-8") as f:
                f.write(f"# {pdf_path.stem}\n\n")
                f.write(f"*Source: {pdf_path.name}*\n\n")
                f.write(markdown_content)

            print(f"  -> Saved to: {output_path}")
        except Exception as e:
            print(f"  -> Error: {e}")


def find_and_convert_excel():
    """Find all Excel files and convert them to JSON."""
    excel_files = list(BASE_DIR.rglob("*.xlsx"))

    # Create output directory
    output_dir = BASE_DIR / "converted" / "json"
    output_dir.mkdir(parents=True, exist_ok=True)

    for excel_path in excel_files:
        # Skip temp files
        if excel_path.name.startswith("~$"):
            continue

        print(f"Converting Excel: {excel_path.name}")
        try:
            json_data = convert_excel_to_json(excel_path)

            # Create output filename
            output_name = excel_path.stem + ".json"
            output_path = output_dir / output_name

            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(json_data, f, indent=2, default=str)

            print(f"  -> Saved to: {output_path}")
        except Exception as e:
            print(f"  -> Error: {e}")


if __name__ == "__main__":
    print("=" * 60)
    print("Converting PDFs to Markdown...")
    print("=" * 60)
    find_and_convert_pdfs()

    print("\n" + "=" * 60)
    print("Converting Excel files to JSON...")
    print("=" * 60)
    find_and_convert_excel()

    print("\n" + "=" * 60)
    print("Done!")
    print("=" * 60)
