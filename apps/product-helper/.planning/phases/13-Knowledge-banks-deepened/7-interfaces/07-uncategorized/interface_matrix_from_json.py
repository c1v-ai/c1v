#!/usr/bin/env python3
"""
interface_matrix_from_json.py — Generate an Interface Matrix workbook from a JSON spec.

Implements Steps 08-09 of the Defining Interfaces KB. Layout matches
`Interface-matrix-sample-basic.xlsx`:

  - One worksheet per subsystem (the OWNER tab).
  - Subsystem ID columns (A, B, C, ...) on the left for `Provided To` markings.
    The owner's own column is left blank (you don't provide to yourself).
  - Then: Value | Units | Estimate? | gap | Last Updated | Last Updated By |
          Interface Champion | gap | Est. Update Due Date | Actual Due Date |
          gap | Row # | Group (Q-col) | Detail (R-col).

Usage:
    python3 interface_matrix_from_json.py <spec.json> [output.xlsx]
    python3 interface_matrix_from_json.py <spec.json> --only <subsystem_id>

The `--only` flag generates a single-tab workbook for incremental review of one
subsystem at a time. Without it, all subsystems become tabs in one workbook.

Validates: subsystem ids unique; every `provided_to` reference exists; owner is
not in its own `provided_to`. Exits non-zero on errors.
"""

from __future__ import annotations

import json
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------- styles

OWNER_GRAY = PatternFill("solid", fgColor="D0D0D0")
HEADER_FILL = PatternFill("solid", fgColor="E8E8E8")
GROUP_FILL = PatternFill("solid", fgColor="F8F8E0")  # subtle highlight for group rows

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT     = Font(name="Arial", size=12, bold=True)
HEADER_FONT    = Font(name="Arial", size=10, bold=True)
GROUP_FONT     = Font(name="Arial", size=10, bold=True)
DETAIL_FONT    = Font(name="Arial", size=10)
INSTR_FONT     = Font(name="Arial", size=10)
INSTR_BOLD     = Font(name="Arial", size=11, bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# --------------------------------------------------------------------------- helpers

def strip_comments(obj):
    if isinstance(obj, dict):
        return {k: strip_comments(v) for k, v in obj.items() if not k.startswith("_")}
    if isinstance(obj, list):
        return [strip_comments(v) for v in obj]
    return obj


def load_spec(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return strip_comments(json.load(f))


def parse_date(value):
    """Accept ISO date strings (YYYY-MM-DD) or pre-parsed datetimes; return datetime or None."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.strptime(str(value), "%Y-%m-%d")
    except ValueError:
        return None


def coerce_value(v):
    """Pass numbers/dates through; if a string starts with '=' Excel will treat as a formula."""
    return v


def validate(spec: dict) -> list[str]:
    errs: list[str] = []
    ss_ids = [s["id"] for s in spec.get("subsystems", [])]
    if len(ss_ids) != len(set(ss_ids)):
        errs.append("Duplicate subsystem ids in `subsystems`.")
    ss_set = set(ss_ids)

    tabs = spec.get("tabs", {})
    for tab_id, rows in tabs.items():
        if tab_id not in ss_set:
            errs.append(f"Tab {tab_id!r} has no matching entry in `subsystems`.")
            continue
        for i, row in enumerate(rows, start=1):
            if "group" not in row and "name" not in row:
                errs.append(f"Tab {tab_id!r}, row {i}: must have `group` or `name`.")
            for target in row.get("provided_to", []) or []:
                if target not in ss_set:
                    errs.append(
                        f"Tab {tab_id!r}, row {i}: provided_to references unknown subsystem {target!r}."
                    )
                if target == tab_id:
                    errs.append(
                        f"Tab {tab_id!r}, row {i}: provided_to cannot include the owner ({tab_id!r})."
                    )
    return errs


# --------------------------------------------------------------------------- layout

def column_layout(n_subsystems: int) -> dict:
    """Compute 1-based column indices given N subsystems on the left.

    Subsystem cols are 1..N. Then a 1-col gap, then Value/Units/Estimate?,
    a gap, Last Updated/By/Champion, a gap, Est. Due/Actual Due, a gap,
    Row #/Group/Detail.
    """
    n = n_subsystems
    cur = n + 2  # gap after subsystem cols
    layout = {
        "value":    cur,
        "units":    cur + 1,
        "estimate": cur + 2,
    }
    cur = layout["estimate"] + 2  # gap
    layout["last_updated"]    = cur
    layout["last_updated_by"] = cur + 1
    layout["champion"]        = cur + 2
    cur = layout["champion"] + 2  # gap
    layout["est_due"]    = cur
    layout["actual_due"] = cur + 1
    cur = layout["actual_due"] + 2  # gap
    layout["row_num"] = cur
    layout["group"]   = cur + 1
    layout["detail"]  = cur + 2
    return layout


# --------------------------------------------------------------------------- writers

def write_instructions_sheet(wb: Workbook, project: dict) -> None:
    ws = wb.create_sheet("Instructions", 0)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 110
    ws["B5"] = project.get("name", "Interface Matrix")
    ws["B5"].font = Font(name="Arial", size=14, bold=True)
    ws["B6"] = f"v{project.get('version','')}  |  {project.get('date','')}"
    ws["B6"].font = INSTR_FONT
    ws["B8"] = "Instructions:"
    ws["B8"].font = INSTR_BOLD
    ws["B9"] = (
        "This workbook contains one tab per subsystem. Each row records an "
        "interface specification owned by that subsystem. `Provided To` columns "
        "(left side) mark which other subsystems consume that spec. The Value "
        "and Units columns are kept separate to prevent unit-conversion errors. "
        "Group headers (column Q) collect related sub-specs (column R). The "
        "Row # column auto-increments via Excel formulas. See "
        "`08 - Creating the Interface Matrix.md` and `09 - Adding Values and Units.md`."
    )
    ws["B9"].font = INSTR_FONT
    ws["B9"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[9].height = 90


def write_subsystem_tab(
    wb: Workbook,
    owner: dict,
    ordered_subsystems: list[dict],
    rows: list[dict],
    starting_row_number: int,
) -> None:
    name_safe = owner["name"][:31]  # Excel sheet name limit
    ws = wb.create_sheet(title=name_safe)

    n = len(ordered_subsystems)
    layout = column_layout(n)
    last_col = layout["detail"]
    last_col_letter = get_column_letter(last_col)

    # ── widths ──
    for j, ss in enumerate(ordered_subsystems, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(15, len(ss["name"]) + 2)
    # gaps
    for k in ("value", "units", "estimate", "last_updated", "last_updated_by",
              "champion", "est_due", "actual_due", "row_num", "group", "detail"):
        col = layout[k]
        widths = {
            "value": 18, "units": 9, "estimate": 10,
            "last_updated": 13, "last_updated_by": 16, "champion": 16,
            "est_due": 20, "actual_due": 16,
            "row_num": 8, "group": 32, "detail": 36,
        }
        ws.column_dimensions[get_column_letter(col)].width = widths[k]
    # narrow the gap columns
    for col in (n + 1, layout["estimate"] + 1, layout["champion"] + 1, layout["actual_due"] + 1):
        ws.column_dimensions[get_column_letter(col)].width = 2

    # ── title ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value=f"Interface Matrix Excerpt : {owner['name']} Sheet")
    c.font = TITLE_FONT
    c.alignment = LEFT

    # ── header row 3 ──
    HEADER_ROW = 3
    # subsystem id columns
    for j, ss in enumerate(ordered_subsystems, start=1):
        cell = ws.cell(row=HEADER_ROW, column=j, value=ss["name"])
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = HEADER_FILL
        cell.border = BORDER
    # spec columns
    headers = [
        ("value", "Value"), ("units", "Units"), ("estimate", "Estimate?"),
        ("last_updated", "Last Updated"), ("last_updated_by", "Last Updated By"),
        ("champion", "Interface Champ."),
        ("est_due", "Est. Update Due Date"), ("actual_due", "Actual Due Date"),
        ("row_num", "Row #"), ("group", owner["name"]), ("detail", ""),
    ]
    for key, label in headers:
        cell = ws.cell(row=HEADER_ROW, column=layout[key], value=label)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = HEADER_FILL
        cell.border = BORDER

    # owner column index for grayout
    owner_col = next(j for j, ss in enumerate(ordered_subsystems, start=1) if ss["id"] == owner["id"])

    # ── data rows ──
    DATA_START = 4
    for i, row in enumerate(rows):
        r = DATA_START + i

        # subsystem provided-to marks
        provided_to = set(row.get("provided_to") or [])
        for j, ss in enumerate(ordered_subsystems, start=1):
            cell = ws.cell(row=r, column=j)
            cell.border = BORDER
            cell.alignment = CENTER
            if j == owner_col:
                cell.fill = OWNER_GRAY
            elif ss["id"] in provided_to:
                cell.value = "Provided To"
                cell.font = DETAIL_FONT

        # value / units / estimate
        if "value" in row:
            v = row["value"]
            cell = ws.cell(row=r, column=layout["value"], value=v)
            cell.alignment = CENTER if isinstance(v, (int, float)) else LEFT
            cell.font = DETAIL_FONT
        if "units" in row:
            cell = ws.cell(row=r, column=layout["units"], value=row["units"])
            cell.alignment = CENTER
            cell.font = DETAIL_FONT
        if row.get("estimate"):
            cell = ws.cell(row=r, column=layout["estimate"], value="X")
            cell.alignment = CENTER
            cell.font = DETAIL_FONT

        # tracking metadata
        for key in ("last_updated", "est_due", "actual_due"):
            d = parse_date(row.get(key))
            if d:
                cell = ws.cell(row=r, column=layout[key], value=d)
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = CENTER
                cell.font = DETAIL_FONT
        for key in ("last_updated_by", "champion"):
            if row.get(key):
                cell = ws.cell(row=r, column=layout[key], value=row[key])
                cell.alignment = CENTER
                cell.font = DETAIL_FONT

        # row number (Excel formula chain)
        if i == 0:
            ws.cell(row=r, column=layout["row_num"], value=starting_row_number)
        else:
            prev = f"{get_column_letter(layout['row_num'])}{r-1}"
            ws.cell(row=r, column=layout["row_num"], value=f"={prev}+1")
        ws.cell(row=r, column=layout["row_num"]).alignment = CENTER
        ws.cell(row=r, column=layout["row_num"]).font = DETAIL_FONT

        # group / detail names
        if "group" in row:
            cell = ws.cell(row=r, column=layout["group"], value=row["group"])
            cell.font = GROUP_FONT
            cell.alignment = LEFT
            cell.fill = GROUP_FILL
        if "name" in row:
            cell = ws.cell(row=r, column=layout["detail"], value=row["name"])
            cell.font = DETAIL_FONT
            cell.alignment = LEFT


def build(spec: dict, only: str | None) -> Workbook:
    wb = Workbook()
    # remove default sheet; we'll create our own
    wb.remove(wb.active)

    project = spec.get("project", {})
    write_instructions_sheet(wb, project)

    subsystems = spec.get("subsystems", [])
    tabs = spec.get("tabs", {})
    starting = int(project.get("starting_row_number") or 1)

    targets = [s for s in subsystems if (only is None or s["id"] == only)]
    for owner in targets:
        rows = tabs.get(owner["id"], [])
        write_subsystem_tab(wb, owner, subsystems, rows, starting)

    return wb


# --------------------------------------------------------------------------- main

def main() -> int:
    args = sys.argv[1:]
    if not args:
        print(__doc__, file=sys.stderr)
        return 2

    in_path = args[0]
    only = None
    out_path: str | None = None

    i = 1
    while i < len(args):
        a = args[i]
        if a == "--only":
            i += 1
            only = args[i]
        elif a.endswith(".xlsx"):
            out_path = a
        else:
            print(f"Unknown argument: {a!r}", file=sys.stderr)
            return 2
        i += 1

    spec = load_spec(in_path)
    errors = validate(spec)
    if errors:
        print("Validation errors:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        return 1

    wb = build(spec, only)

    if out_path is None:
        pid = spec.get("project", {}).get("id", "interface-matrix")
        suffix = f"__{only}" if only else ""
        out_path = os.path.join(
            os.path.dirname(os.path.abspath(in_path)),
            f"{pid}{suffix}.xlsx",
        )
    wb.save(out_path)

    n_tabs = len(wb.sheetnames) - 1  # minus Instructions
    print(f"Wrote: {out_path}")
    print(f"  Subsystems in spec: {len(spec.get('subsystems', []))}")
    print(f"  Tabs written: {n_tabs}{' (filter: ' + only + ')' if only else ''}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
