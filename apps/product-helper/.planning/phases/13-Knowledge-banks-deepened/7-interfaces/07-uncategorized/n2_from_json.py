#!/usr/bin/env python3
"""
n2_from_json.py — Generate an N-Squared (N2) Interface Chart from a JSON spec.

Usage:
    python n2_from_json.py <input.json> [output.xlsx]

If output.xlsx is omitted, writes <project.id>-n2-chart.xlsx in the current dir.

The JSON schema is documented in `n2-chart-template.json`. This script:
  1. Validates the spec (subsystem ids unique, interface (from,to) refs valid).
  2. Builds the N2 matrix sheet (diagonal = subsystems, cells = interfaces).
  3. Generates an Analysis sheet with:
        - Critical-subsystem fan-in/fan-out counts (flagged if >50% of N-1).
        - System flows and control loops (validated against the matrix).
        - QFD endpoint reconciliation (sum of endpoints_count vs. target).
  4. Writes an External Interfaces section.

Exits non-zero on validation errors.
"""

from __future__ import annotations

import json
import os
import sys
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------- styles

CATEGORY_FILLS = {
    "engine":   PatternFill("solid", fgColor="FFE0E0"),
    "ui":       PatternFill("solid", fgColor="E0E0FF"),
    "platform": PatternFill("solid", fgColor="E0FFE0"),
    "data":     PatternFill("solid", fgColor="FFF3E0"),
    "external": PatternFill("solid", fgColor="F0F0F0"),
}
DEFAULT_FILL  = PatternFill("solid", fgColor="F5F5F5")
HEADER_FILL   = PatternFill("solid", fgColor="D9D9D9")
EMPTY_FILL    = PatternFill("solid", fgColor="FFFFFF")
LOOP_FILL     = PatternFill("solid", fgColor="FFF8C0")
CRITICAL_FILL = PatternFill("solid", fgColor="FFD0D0")

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT    = Font(name="Arial", size=14, bold=True)
SUBTITLE_FONT = Font(name="Arial", size=10)
SECTION_FONT  = Font(name="Arial", size=11, bold=True)
HEADER_FONT   = Font(name="Arial", size=10, bold=True)
DIAG_FONT     = Font(name="Arial", size=11, bold=True)
CELL_FONT     = Font(name="Arial", size=8)
NOTE_FONT     = Font(name="Arial", size=9)
NOTE_BOLD     = Font(name="Arial", size=9, bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)


# --------------------------------------------------------------------------- model

@dataclass
class Spec:
    project: dict
    subsystems: list[dict]
    interfaces: list[dict]
    externals: list[dict]
    flows: list[dict]
    loops: list[dict]

    @property
    def ss_index(self) -> dict[str, int]:
        return {s["id"]: i for i, s in enumerate(self.subsystems)}


def strip_comments(obj):
    """Recursively remove keys starting with `_` (used as inline comments)."""
    if isinstance(obj, dict):
        return {k: strip_comments(v) for k, v in obj.items() if not k.startswith("_")}
    if isinstance(obj, list):
        return [strip_comments(v) for v in obj]
    return obj


def load_spec(path: str) -> Spec:
    with open(path, encoding="utf-8") as f:
        raw = json.load(f)
    raw = strip_comments(raw)
    return Spec(
        project=raw.get("project", {}),
        subsystems=raw.get("subsystems", []),
        interfaces=raw.get("interfaces", []),
        externals=raw.get("external_interfaces", []),
        flows=raw.get("system_flows", []),
        loops=raw.get("control_loops", []),
    )


def validate(spec: Spec) -> list[str]:
    errs: list[str] = []
    ids = [s["id"] for s in spec.subsystems]
    if len(ids) != len(set(ids)):
        errs.append("Duplicate subsystem ids in `subsystems`.")
    ss_set = set(ids)

    seen_pairs: set[tuple[str, str]] = set()
    seen_ifids: set[str] = set()
    for it in spec.interfaces:
        if it["from"] not in ss_set:
            errs.append(f"Interface {it.get('id', '?')}: unknown `from` {it['from']!r}.")
        if it["to"] not in ss_set:
            errs.append(f"Interface {it.get('id', '?')}: unknown `to` {it['to']!r}.")
        if it["from"] == it["to"]:
            errs.append(f"Interface {it.get('id', '?')}: `from` and `to` are the same ({it['from']}).")
        pair = (it["from"], it["to"])
        if pair in seen_pairs:
            errs.append(f"Duplicate interface pair {pair} (only one entry per direction allowed).")
        seen_pairs.add(pair)
        if it["id"] in seen_ifids:
            errs.append(f"Duplicate interface id {it['id']!r}.")
        seen_ifids.add(it["id"])

    pair_lookup = {(it["from"], it["to"]) for it in spec.interfaces}
    for f in spec.flows:
        path = f.get("path", [])
        for a, b in zip(path, path[1:]):
            if (a, b) not in pair_lookup:
                errs.append(f"Flow {f.get('name', '?')!r}: hop {a} -> {b} has no matching interface.")
    for lp in spec.loops:
        path = lp.get("path", [])
        if len(path) < 3 or path[0] != path[-1]:
            errs.append(f"Loop {lp.get('name', '?')!r}: path must start and end at the same subsystem (length >= 3).")
        for a, b in zip(path, path[1:]):
            if (a, b) not in pair_lookup:
                errs.append(f"Loop {lp.get('name', '?')!r}: hop {a} -> {b} has no matching interface.")
    return errs


# --------------------------------------------------------------------------- analysis

def fan_counts(spec: Spec) -> tuple[dict[str, int], dict[str, int]]:
    out = {s["id"]: 0 for s in spec.subsystems}
    inn = {s["id"]: 0 for s in spec.subsystems}
    for it in spec.interfaces:
        out[it["from"]] += 1
        inn[it["to"]]   += 1
    return out, inn


def critical_threshold(n: int) -> int:
    """Critical = fan-in or fan-out > 50% of (n-1)."""
    return max(1, (n - 1) // 2 + 1)


def endpoint_reconciliation(spec: Spec) -> tuple[int, int, float]:
    declared = sum(int(it.get("endpoints_count") or 0) for it in spec.interfaces)
    target   = int(spec.project.get("qfd_target_endpoints") or 0)
    drift    = 0.0 if target == 0 else (declared - target) / target * 100.0
    return declared, target, drift


# --------------------------------------------------------------------------- writer

def fill_for(category: str | None) -> PatternFill:
    return CATEGORY_FILLS.get((category or "").lower(), DEFAULT_FILL)


def cell_set(ws, row, col, value, *, font=None, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.border = BORDER
    if font:  c.font = font
    if fill:  c.fill = fill
    if align: c.alignment = align
    return c


def write_n2_sheet(wb: Workbook, spec: Spec, loop_pairs: set[tuple[int, int]]) -> None:
    ws = wb.active
    ws.title = "N2 Chart"

    n = len(spec.subsystems)
    iface = {(spec.ss_index[i["from"]], spec.ss_index[i["to"]]): i for i in spec.interfaces}

    # widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18
    for j in range(n):
        ws.column_dimensions[get_column_letter(j + 3)].width = 22

    # title
    last_col_letter = get_column_letter(n + 2)
    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws["A1"]
    c.value = spec.project.get("name", "N2 Interface Chart")
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A2:{last_col_letter}2")
    declared, target, drift = endpoint_reconciliation(spec)
    subtitle = (
        f"Type: {spec.project.get('interface_type', '?')}  |  "
        f"v{spec.project.get('version', '?')}  |  {spec.project.get('date', '')}  |  "
        f"{len(spec.interfaces)} internal + {len(spec.externals)} external interfaces"
    )
    if target:
        subtitle += f"  |  endpoints: {declared} declared vs. {target} QFD target ({drift:+.0f}%)"
    c2 = ws["A2"]
    c2.value = subtitle
    c2.font = SUBTITLE_FONT

    # header row
    header_row = 4
    cell_set(ws, header_row, 1, "", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    cell_set(ws, header_row, 2, "FROM \\ TO", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    for j, ss in enumerate(spec.subsystems):
        cell_set(ws, header_row, j + 3,
                 f"{ss['id']}\n{ss['name']}",
                 font=HEADER_FONT, fill=fill_for(ss.get("category")), align=CENTER)
    ws.row_dimensions[header_row].height = 45

    # critical-subsystem flagging
    out_cnt, in_cnt = fan_counts(spec)
    threshold = critical_threshold(n)

    # body rows
    for i, ss in enumerate(spec.subsystems):
        row = header_row + 1 + i
        ws.row_dimensions[row].height = 70
        cell_set(ws, row, 1, i + 1, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        critical = (out_cnt[ss["id"]] >= threshold) or (in_cnt[ss["id"]] >= threshold)
        row_header_fill = CRITICAL_FILL if critical else fill_for(ss.get("category"))
        cell_set(ws, row, 2, f"{ss['id']}\n{ss['name']}",
                 font=HEADER_FONT, fill=row_header_fill, align=CENTER)

        for j in range(n):
            col = j + 3
            if i == j:
                cell_set(ws, row, col, f"{ss['id']}\n{ss['name']}",
                         font=DIAG_FONT, fill=fill_for(ss.get("category")), align=CENTER)
            elif (i, j) in iface:
                it = iface[(i, j)]
                txt = f"{it['id']}\n{it.get('name', '')}\n{it.get('data', '')}".strip()
                fill = LOOP_FILL if (i, j) in loop_pairs else EMPTY_FILL
                cell_set(ws, row, col, txt, font=CELL_FONT, fill=fill, align=WRAP)
            else:
                cell_set(ws, row, col, "", fill=EMPTY_FILL, align=WRAP)

    # external interfaces section
    if spec.externals:
        sec_row = header_row + 1 + n + 1
        ws.merge_cells(f"A{sec_row}:{last_col_letter}{sec_row}")
        ws.cell(row=sec_row, column=1, value="External Interfaces").font = SECTION_FONT

        hdr = sec_row + 1
        for c_idx, label in enumerate(["IF-ID", "From", "To", "Data / Signal"], start=1):
            cell_set(ws, hdr, c_idx, label, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        ws.merge_cells(f"D{hdr}:{last_col_letter}{hdr}")

        for k, ext in enumerate(spec.externals):
            r = hdr + 1 + k
            ws.row_dimensions[r].height = 30
            cell_set(ws, r, 1, ext.get("id", ""), font=CELL_FONT, fill=fill_for("external"), align=CENTER)
            cell_set(ws, r, 2, ext.get("from", ""), font=CELL_FONT, align=CENTER)
            cell_set(ws, r, 3, ext.get("to", ""),   font=CELL_FONT, align=CENTER)
            ws.merge_cells(f"D{r}:{last_col_letter}{r}")
            cell_set(ws, r, 4, ext.get("data", ""), font=CELL_FONT, align=WRAP)


def write_analysis_sheet(wb: Workbook, spec: Spec) -> set[tuple[int, int]]:
    """Write Analysis sheet. Returns the set of (i,j) cells that are in any control loop."""
    ws = wb.create_sheet("Analysis")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 18

    n = len(spec.subsystems)
    out_cnt, in_cnt = fan_counts(spec)
    threshold = critical_threshold(n)

    row = 1
    ws.cell(row=row, column=1, value="N2 Chart — Analysis").font = TITLE_FONT
    row += 2

    # critical subsystems
    ws.cell(row=row, column=1, value="Critical Subsystems (fan-in or fan-out >= 50% of N-1)").font = SECTION_FONT
    row += 1
    for col, label in enumerate(["Subsystem", "Outputs (provides to)", "Inputs (receives from)"], start=1):
        cell_set(ws, row, col, label, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    row += 1
    for ss in spec.subsystems:
        critical = (out_cnt[ss["id"]] >= threshold) or (in_cnt[ss["id"]] >= threshold)
        flag = "  CRITICAL" if critical else ""
        cell_set(ws, row, 1, f"{ss['id']} {ss['name']}{flag}",
                 font=NOTE_BOLD if critical else NOTE_FONT,
                 fill=CRITICAL_FILL if critical else fill_for(ss.get("category")))
        cell_set(ws, row, 2, str(out_cnt[ss["id"]]), font=NOTE_FONT, align=CENTER)
        cell_set(ws, row, 3, str(in_cnt[ss["id"]]), font=NOTE_FONT, align=CENTER)
        row += 1
    row += 1

    # system flows
    if spec.flows:
        ws.cell(row=row, column=1, value="System Flows").font = SECTION_FONT
        row += 1
        for f in spec.flows:
            cell_set(ws, row, 1, f.get("name", ""), font=NOTE_BOLD)
            cell_set(ws, row, 2, " -> ".join(f.get("path", [])) + "\n" + f.get("description", ""),
                     font=NOTE_FONT, align=WRAP)
            ws.row_dimensions[row].height = 40
            row += 1
        row += 1

    # control loops
    loop_pairs: set[tuple[int, int]] = set()
    if spec.loops:
        ws.cell(row=row, column=1, value="Control Loops (highlighted yellow on N2 sheet)").font = SECTION_FONT
        row += 1
        for lp in spec.loops:
            path = lp.get("path", [])
            for a, b in zip(path, path[1:]):
                loop_pairs.add((spec.ss_index[a], spec.ss_index[b]))
            cell_set(ws, row, 1, lp.get("name", ""), font=NOTE_BOLD, fill=LOOP_FILL)
            cell_set(ws, row, 2, " -> ".join(path) + "\n" + lp.get("description", ""),
                     font=NOTE_FONT, align=WRAP)
            ws.row_dimensions[row].height = 40
            row += 1
        row += 1

    # QFD reconciliation
    declared, target, drift = endpoint_reconciliation(spec)
    if target:
        ws.cell(row=row, column=1, value="QFD Endpoint Reconciliation").font = SECTION_FONT
        row += 1
        verdict = "ALIGNED" if abs(drift) <= 25 else "DRIFT — investigate"
        for col, (label, value) in enumerate([
            ("QFD design target",    str(target)),
            ("Declared in this N2",  str(declared)),
            ("Drift",                f"{drift:+.0f}%  ({verdict})"),
        ], start=0):
            cell_set(ws, row + col, 1, label, font=NOTE_BOLD)
            cell_set(ws, row + col, 2, value, font=NOTE_FONT)
        row += 4

    return loop_pairs


# --------------------------------------------------------------------------- main

def build(spec: Spec, out_path: str) -> None:
    wb = Workbook()
    # Analysis sheet computes loop_pairs; build it first into a temp, then move.
    # We need loop_pairs before writing N2, so: build analysis on a side workbook? No —
    # simpler: compute loop_pairs directly, then write both sheets.
    loop_pairs: set[tuple[int, int]] = set()
    for lp in spec.loops:
        path = lp.get("path", [])
        for a, b in zip(path, path[1:]):
            loop_pairs.add((spec.ss_index[a], spec.ss_index[b]))

    write_n2_sheet(wb, spec, loop_pairs)
    write_analysis_sheet(wb, spec)
    wb.save(out_path)


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__, file=sys.stderr)
        return 2
    in_path = sys.argv[1]
    spec = load_spec(in_path)

    errors = validate(spec)
    if errors:
        print("Validation errors:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        return 1

    if len(sys.argv) >= 3:
        out_path = sys.argv[2]
    else:
        pid = spec.project.get("id", "n2-chart")
        out_path = os.path.join(os.path.dirname(os.path.abspath(in_path)), f"{pid}-n2-chart.xlsx")

    build(spec, out_path)

    declared, target, drift = endpoint_reconciliation(spec)
    print(f"Wrote: {out_path}")
    print(f"  Subsystems: {len(spec.subsystems)}")
    print(f"  Interfaces: {len(spec.interfaces)} internal + {len(spec.externals)} external")
    print(f"  Flows: {len(spec.flows)}  Control loops: {len(spec.loops)}")
    if target:
        print(f"  QFD endpoints: {declared} declared vs. {target} target ({drift:+.0f}%)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
