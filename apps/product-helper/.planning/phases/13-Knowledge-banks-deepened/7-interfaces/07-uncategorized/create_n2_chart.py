#!/usr/bin/env python3
"""
Generate THG N2 Interface Chart — Option B (7 subsystems).

Adds SS7 Team Coordination Engine to the original 6 subsystems.
Outputs: THG_N2_Chart_v2.xlsx

N2 chart methodology (CESYS526):
  - Subsystems on the diagonal
  - Row = FROM (provider), Column = TO (receiver)
  - Each off-diagonal cell shows interface ID + data description
  - System flows, control loops, and critical subsystems identified

Original: 6 subsystems, 13 internal + 4 external = 17 interfaces
Option B: 7 subsystems, 21 internal + 4 external = 25 interfaces
  New interfaces: IF-18 through IF-25 (8 new)
    IF-25: SS2 -> SS1 intervention outcomes for ML training (patent feedback loop)
  Modified interfaces: IF-01, IF-02, IF-03, IF-04, IF-05, IF-06, IF-07,
                       IF-08, IF-09, IF-12, IF-14 (10 of 17 updated)
  Patent sub-modules: F.4a Stratification (in SS1), F.5.4a Intervention (in SS2)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers,
)
from openpyxl.utils import get_column_letter
import os

# ── Colors ──────────────────────────────────────────────────────
PRED_FILL = PatternFill("solid", fgColor="FFE0E0")   # Light red — engines
ALERT_FILL = PatternFill("solid", fgColor="FFE0E0")
DASH_FILL = PatternFill("solid", fgColor="E0E0FF")   # Light blue — UI
MOBILE_FILL = PatternFill("solid", fgColor="E0E0FF")
PLATFORM_FILL = PatternFill("solid", fgColor="E0FFE0")  # Light green — platform
REG_FILL = PatternFill("solid", fgColor="E0FFE0")
TEAM_FILL = PatternFill("solid", fgColor="FFF3E0")    # Light orange — NEW patent
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")  # Gray header
EMPTY_FILL = PatternFill("solid", fgColor="FFFFFF")
NEW_IF_FILL = PatternFill("solid", fgColor="FFFFF0")  # Light yellow — new interfaces
MODIFIED_FILL = PatternFill("solid", fgColor="FFF8F0") # Very light orange — modified
EXT_FILL = PatternFill("solid", fgColor="F0F0F0")     # Gray — external

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FONT = Font(name="Arial", size=10, bold=True)
DIAG_FONT = Font(name="Arial", size=11, bold=True)
CELL_FONT = Font(name="Arial", size=8)
CELL_FONT_NEW = Font(name="Arial", size=8, color="8B4513")  # Brown for new
TITLE_FONT = Font(name="Arial", size=14, bold=True)
SUBTITLE_FONT = Font(name="Arial", size=10)
SECTION_FONT = Font(name="Arial", size=11, bold=True)
NOTE_FONT = Font(name="Arial", size=9)
NOTE_FONT_BOLD = Font(name="Arial", size=9, bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
WRAP_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

# ── Subsystem Definitions ───────────────────────────────────────
SUBSYSTEMS = [
    ("SS1", "Prediction\nEngine", PRED_FILL),
    ("SS2", "Alert\nEngine", ALERT_FILL),
    ("SS3", "Dashboard", DASH_FILL),
    ("SS4", "Mobile\nApp", MOBILE_FILL),
    ("SS5", "Platform\nCore", PLATFORM_FILL),
    ("SS6", "Regulatory\nEngine", REG_FILL),
    ("SS7", "Team Coord\nEngine", TEAM_FILL),
]

N = len(SUBSYSTEMS)

# ── Interface Matrix ────────────────────────────────────────────
# Key: (from_idx, to_idx) -> (IF_ID, description, is_new, is_modified)
# from_idx/to_idx are 0-based indices into SUBSYSTEMS
# is_new = True for IF-18 through IF-24
# is_modified = True for interfaces updated by DBML extension

INTERFACES = {
    # SS1 Prediction Engine outputs
    (0, 1): ("IF-02", "Prediction result:\nTrec, Dlim, risk_level,\ncompliance_trigger,\nwbgt_effective", False, True),
    (0, 4): ("IF-06", "Prediction audit log:\ninputs, outputs, timestamps,\nwbgt_screening_detail,\nml_model_id, acclim day", False, True),
    (0, 6): ("IF-18", "Member prediction state:\nrisk_level, dlim_minutes,\nrecovery_rate_score\nper team member", True, False),

    # SS2 Alert Engine outputs
    (1, 0): ("IF-25", "Intervention outcomes:\neffectiveness_score,\ntrec_reduction, dlim_recovery\n-> ML training pipeline", True, False),
    (1, 2): ("IF-04", "Alert lifecycle:\nnew alerts, acks, escalations,\nintervention outcomes,\neffectiveness scores", False, True),
    (1, 3): ("IF-03", "Alert + prescription:\nrisk level, guidance,\nfluid_vol, cooling_method,\nrest_duration, equip_removal", False, True),
    (1, 4): ("IF-07", "Alert lifecycle audit:\ntrigger -> prescription ->\nack -> escalation ->\nintervention -> resolution", False, True),

    # SS3 Dashboard outputs
    (2, 5): ("IF-13", "Report request:\nsite ID, date range,\nreport type", False, False),
    (2, 6): ("IF-20", "Manager team commands:\ncreate/end session,\naccept/reject rotation,\nassign workers", True, False),

    # SS4 Mobile App outputs
    (3, 0): ("IF-01", "Worker inputs:\nactivity, clothing/PPE,\nprofile ID, surface_type,\nactivity switch data", False, True),
    (3, 1): ("IF-05", "Acknowledgment:\nworker ID, alert ID,\nGPS, timestamp,\npre_intervention_trec", False, True),

    # SS5 Platform Core outputs
    (4, 0): ("IF-12", "Tenant + worker config:\ntenant config, worker profiles,\nsites, acclimatization data,\nregulatory_standard per site", False, True),
    (4, 2): ("IF-11", "Auth context:\nJWT token, RBAC perms,\ntenant ID,\ndashboard scope", False, False),
    (4, 3): ("IF-10", "Auth context:\nJWT token, RBAC perms,\ntenant ID", False, False),
    (4, 6): ("IF-24", "Auth context:\nJWT token, RBAC perms,\ntenant ID\nfor manager actions", True, False),

    # SS6 Regulatory Engine outputs
    (5, 0): ("IF-23", "WBGT thresholds:\nmetabolic x regimen x acclim\n-> AL/TLV limits\nfor dual-model screening", True, False),
    (5, 1): ("IF-08", "Active thresholds:\nrisk levels, escalation times,\nWBGT thresholds per\nstate/org/metabolic/regimen", False, True),
    (5, 2): ("IF-09", "Report templates:\njurisdiction-specific formats,\nmulti-state resolution", False, True),

    # SS7 Team Coordination Engine outputs
    (6, 2): ("IF-19", "Team session state:\nsessions, member risk,\npending rotations,\npriority levels", True, False),
    (6, 3): ("IF-21", "Worker rotation notify:\nrotation accepted,\nnew assignment,\naffected worker IDs", True, False),
    (6, 4): ("IF-22", "Team coordination audit:\nsession lifecycle,\nrotation decisions,\nmember join/leave", True, False),
}

# ── External Interfaces ─────────────────────────────────────────
EXTERNALS = [
    ("IF-14", "Weather API", "SS1", "Env data: temp, humidity, wind,\nsolar, dew_point, pressure_hpa", True),
    ("IF-15", "SSO Provider", "SS5", "SAML assertion:\nauthenticated identity, groups", False),
    ("IF-16", "SS2", "Emergency\nServices", "Critical alert: worker ID,\nlocation, risk level", False),
    ("IF-17", "SS5", "Audit\nSystems", "Compliance export:\nimmutable audit logs", False),
]


def create_n2_sheet(wb):
    ws = wb.active
    ws.title = "N2 Chart"

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 16
    for i in range(N):
        col = get_column_letter(i + 3)  # C through I
        ws.column_dimensions[col].width = 22

    # ── Title rows ──
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "Team Heat Guard -- N2 Interface Chart (Option B: 7 Subsystems)"
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = "CESYS526 | 21 internal + 4 external = 25 interfaces | Updated 2026-04-12"
    c.font = SUBTITLE_FONT

    # ── Column headers (TO subsystems) ──
    row = 4
    ws.cell(row=row, column=1, value="").border = THIN_BORDER
    ws.cell(row=row, column=2, value="FROM \\ TO").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).alignment = CENTER
    ws.cell(row=row, column=2).border = THIN_BORDER

    for j, (ss_id, ss_name, fill) in enumerate(SUBSYSTEMS):
        col = j + 3
        cell = ws.cell(row=row, column=col)
        cell.value = f"{ss_id}\n{ss_name}"
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 45

    # ── Matrix rows ──
    for i, (ss_id, ss_name, fill) in enumerate(SUBSYSTEMS):
        row = i + 5
        ws.row_dimensions[row].height = 80

        # Row number
        cell = ws.cell(row=row, column=1)
        cell.value = i + 1
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = HEADER_FILL

        # Row header (FROM subsystem)
        cell = ws.cell(row=row, column=2)
        cell.value = f"{ss_id}\n{ss_name}"
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER

        # Matrix cells
        for j in range(N):
            col = j + 3
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = WRAP

            if i == j:
                # Diagonal cell = subsystem name
                cell.value = f"{SUBSYSTEMS[i][0]}\n{SUBSYSTEMS[i][1]}"
                cell.font = DIAG_FONT
                cell.fill = SUBSYSTEMS[i][2]
                cell.alignment = CENTER
            elif (i, j) in INTERFACES:
                if_id, desc, is_new, is_modified = INTERFACES[(i, j)]
                cell.value = f"{if_id}\n{desc}"
                if is_new:
                    cell.font = CELL_FONT_NEW
                    cell.fill = NEW_IF_FILL
                elif is_modified:
                    cell.font = CELL_FONT
                    cell.fill = MODIFIED_FILL
                else:
                    cell.font = CELL_FONT
                    cell.fill = EMPTY_FILL
            else:
                cell.value = ""
                cell.fill = EMPTY_FILL

    # ── External interfaces section ──
    ext_start = 5 + N + 1  # row 13
    ws.merge_cells(f"A{ext_start}:I{ext_start}")
    cell = ws.cell(row=ext_start, column=1)
    cell.value = "External Interfaces"
    cell.font = SECTION_FONT

    ext_row = ext_start + 1
    headers = ["IF-ID", "From", "To", "Data / Signal", "Modified?"]
    header_cols = [1, 2, 3, 4, 7]
    for h, c_idx in zip(headers, header_cols):
        cell = ws.cell(row=ext_row, column=c_idx)
        cell.value = h
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    # Merge data column
    ws.merge_cells(f"D{ext_row}:F{ext_row}")
    ws.merge_cells(f"G{ext_row}:I{ext_row}")

    for idx, (if_id, frm, to, data, modified) in enumerate(EXTERNALS):
        r = ext_row + 1 + idx
        ws.row_dimensions[r].height = 40

        ws.cell(row=r, column=1, value=if_id).font = CELL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).fill = EXT_FILL

        ws.cell(row=r, column=2, value=frm).font = CELL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).alignment = CENTER

        ws.cell(row=r, column=3, value=to).font = CELL_FONT
        ws.cell(row=r, column=3).border = THIN_BORDER
        ws.cell(row=r, column=3).alignment = CENTER

        ws.merge_cells(f"D{r}:F{r}")
        ws.cell(row=r, column=4, value=data).font = CELL_FONT
        ws.cell(row=r, column=4).border = THIN_BORDER
        ws.cell(row=r, column=4).alignment = WRAP

        ws.merge_cells(f"G{r}:I{r}")
        ws.cell(row=r, column=7, value="Yes" if modified else "").font = CELL_FONT
        ws.cell(row=r, column=7).border = THIN_BORDER
        ws.cell(row=r, column=7).alignment = CENTER
        if modified:
            ws.cell(row=r, column=7).fill = MODIFIED_FILL

    return ext_row + len(EXTERNALS) + 1


def create_analysis_sheet(wb, after_ext_row):
    ws = wb.create_sheet("Analysis")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20

    row = 1
    ws.cell(row=row, column=1, value="N2 Chart Analysis -- Option B").font = TITLE_FONT
    row += 2

    # ── System Flows ──
    ws.cell(row=row, column=1, value="System Flows Identified").font = SECTION_FONT
    row += 1

    flows = [
        ("1. Prediction Loop (primary)",
         "Weather API -> SS1 -> SS2 -> SS4 -> Worker\n"
         "Env data -> dual-model prediction -> classification + intervention -> alert delivery",
         "CRITICAL PATH"),
        ("2. Escalation Flow",
         "SS2 -> SS4 (no ack) -> SS2 -> Supervisor -> SS5(audit)\n"
         "Alert sent -> timeout (AMBER 5min/RED 2min) -> escalation -> re-escalation 10min",
         "Safety-critical"),
        ("3. Compliance Flow",
         "SS3 -> SS6 -> SS3 -> Safety Manager\n"
         "Report request -> jurisdiction template resolution -> compiled report",
         "Regulatory"),
        ("4. Control Loop",
         "SS1 -> SS2 -> SS4 -> SS2 -> SS1 (repeat)\n"
         "Predict -> classify -> alert -> acknowledge -> re-predict next cycle",
         "Continuous"),
        ("5. Team Coordination Loop (NEW)",
         "SS1 -> SS7 -> SS3 -> SS7 -> SS4\n"
         "Prediction updates member state -> rotation detected -> manager reviews ->\n"
         "accept/reject -> notify worker",
         "PATENT"),
        ("6. Intervention Feedback Loop (NEW)",
         "SS2 -> SS4 -> SS2 -> SS1 -> SS2\n"
         "Prescription -> worker completes -> record outcome + effectiveness ->\n"
         "re-predict -> new risk level",
         "PATENT (ML training)"),
    ]

    for name, desc, category in flows:
        ws.cell(row=row, column=1, value=name).font = NOTE_FONT_BOLD
        ws.cell(row=row, column=2, value=desc).font = NOTE_FONT
        ws.cell(row=row, column=2).alignment = WRAP
        ws.cell(row=row, column=3, value=category).font = NOTE_FONT
        ws.row_dimensions[row].height = 55
        row += 1

    row += 1

    # ── Critical Subsystem Analysis ──
    ws.cell(row=row, column=1, value="Critical Subsystem Analysis").font = SECTION_FONT
    row += 1

    # Count inputs/outputs per subsystem
    output_counts = {}
    input_counts = {}
    for ss_id, _, _ in SUBSYSTEMS:
        output_counts[ss_id] = 0
        input_counts[ss_id] = 0

    for (i, j), (if_id, _, _, _) in INTERFACES.items():
        output_counts[SUBSYSTEMS[i][0]] += 1
        input_counts[SUBSYSTEMS[j][0]] += 1

    # Add external
    output_counts["SS2"] += 1  # IF-16
    output_counts["SS5"] += 1  # IF-17
    input_counts["SS1"] += 1   # IF-14
    input_counts["SS5"] += 1   # IF-15

    ws.cell(row=row, column=1, value="Subsystem").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Outputs (provides to)").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=3, value="Inputs (receives from)").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = HEADER_FILL
    row += 1

    for ss_id, ss_name, fill in SUBSYSTEMS:
        name_clean = ss_name.replace("\n", " ")
        ws.cell(row=row, column=1, value=f"{ss_id} {name_clean}").font = NOTE_FONT
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=2, value=output_counts[ss_id]).font = NOTE_FONT
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=3, value=input_counts[ss_id]).font = NOTE_FONT
        ws.cell(row=row, column=3).alignment = CENTER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Key Findings:").font = SECTION_FONT
    row += 1

    findings = [
        "SS5 (Platform Core): Most outputs (4) -- auth/config hub. Single point of failure.",
        "SS2 (Alert Engine): 4 outputs, 3 inputs -- highest traffic. Now feeds ML training back to SS1 (IF-25).",
        "SS1 (Prediction Engine): Feeds SS2 + SS7 + SS5. Receives from SS2 (IF-25 ML feedback). Core value.",
        "SS7 (Team Coordination): 3 inputs, 3 outputs. Well-connected but not overloaded.",
        "SS6 -> SS1 (IF-23): NEW dependency. SS1 needs WBGT thresholds directly for dual-model screening.",
        "SS2 -> SS1 (IF-25): NEW feedback loop. Intervention outcomes train ML models in SS1.",
        "Patent features: F.4a Stratification (sub-module of SS1), F.5.4a Intervention (sub-module of SS2).",
    ]

    for f in findings:
        ws.cell(row=row, column=1, value=f).font = NOTE_FONT
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row=row, column=1).alignment = WRAP
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # ── Change Summary ──
    ws.cell(row=row, column=1, value="Change Summary vs Original N2").font = SECTION_FONT
    row += 1

    changes = [
        ("Original", "6 subsystems, 13 internal + 4 external = 17 total"),
        ("Option B", "7 subsystems, 21 internal + 4 external = 25 total"),
        ("New subsystem", "SS7 Team Coordination Engine (PATENT)"),
        ("New interfaces (8)", "IF-18 (SS1->SS7), IF-19 (SS7->SS3), IF-20 (SS3->SS7),\n"
         "IF-21 (SS7->SS4), IF-22 (SS7->SS5), IF-23 (SS6->SS1),\n"
         "IF-24 (SS5->SS7), IF-25 (SS2->SS1 intervention ML feedback)"),
        ("Modified interfaces (10)", "IF-01, IF-02, IF-03, IF-04, IF-05, IF-06, IF-07,\n"
         "IF-08, IF-09, IF-12, IF-14"),
        ("Unchanged (6)", "IF-10, IF-11, IF-13, IF-15, IF-16, IF-17"),
        ("Patent sub-modules", "F.4a Pre-Event Stratification visible in SS1\n"
         "F.5.4a AI Intervention Engine visible in SS2"),
    ]

    for label, desc in changes:
        ws.cell(row=row, column=1, value=label).font = NOTE_FONT_BOLD
        ws.cell(row=row, column=2, value=desc).font = NOTE_FONT
        ws.cell(row=row, column=2).alignment = WRAP
        ws.row_dimensions[row].height = 35
        row += 1

    row += 2

    # ── Requirement Allocation to SS7 ──
    ws.cell(row=row, column=1, value="SS7 Requirement Allocation").font = SECTION_FONT
    row += 1

    alloc = [
        ("TEAM-01 to TEAM-11", "Primary owner: SS7 Team Coordination Engine"),
        ("DASH-08, DASH-09, DASH-10, DASH-14", "Display: SS3 | Logic: SS7"),
        ("STRT-01 to STRT-09", "Stays in SS1 (uses prediction models, forward-looking PHS)"),
        ("ESCL-01 to ESCL-06", "Stays in SS2 (extension of alert escalation)"),
        ("All other DBML tables", "ML infra (C.1-C.2) -> SS1 | Intervention (G.1-G.2) -> SS2"),
    ]

    for reqs, owner in alloc:
        ws.cell(row=row, column=1, value=reqs).font = NOTE_FONT_BOLD
        ws.cell(row=row, column=2, value=owner).font = NOTE_FONT
        ws.cell(row=row, column=2).alignment = WRAP
        ws.row_dimensions[row].height = 25
        row += 1


def create_legend_sheet(wb):
    ws = wb.create_sheet("Legend")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    row = 1
    ws.cell(row=row, column=1, value="N2 Chart Legend").font = TITLE_FONT
    row += 2

    legend = [
        ("Diagonal cells", "Subsystem names (colored by category)"),
        ("Off-diagonal cells", "Interface FROM row subsystem TO column subsystem"),
        ("Light yellow cells", "NEW interfaces added in Option B (IF-18 to IF-25)"),
        ("Light orange cells", "MODIFIED interfaces (data expanded by DBML extension)"),
        ("White cells", "UNCHANGED interfaces from original N2"),
        ("Empty cells", "No direct interface between those subsystems"),
        ("Brown text", "New interface text (IF-18 to IF-25)"),
    ]

    for label, desc in legend:
        ws.cell(row=row, column=1, value=label).font = NOTE_FONT_BOLD
        ws.cell(row=row, column=2, value=desc).font = NOTE_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Color Categories").font = SECTION_FONT
    row += 1

    colors = [
        ("Red (light)", "Engine subsystems (SS1, SS2)", PRED_FILL),
        ("Blue (light)", "UI subsystems (SS3, SS4)", DASH_FILL),
        ("Green (light)", "Platform subsystems (SS5, SS6)", PLATFORM_FILL),
        ("Orange (light)", "NEW: Team Coordination (SS7)", TEAM_FILL),
    ]

    for label, desc, fill in colors:
        ws.cell(row=row, column=1, value=label).font = NOTE_FONT
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=2, value=desc).font = NOTE_FONT
        row += 1


if __name__ == "__main__":
    wb = Workbook()

    after_ext = create_n2_sheet(wb)
    create_analysis_sheet(wb, after_ext)
    create_legend_sheet(wb)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    out = os.path.join(script_dir, "THG_N2_Chart_v2.xlsx")
    wb.save(out)

    # Count stats
    new_count = sum(1 for v in INTERFACES.values() if v[2])
    mod_count = sum(1 for v in INTERFACES.values() if v[3])
    unchanged = len(INTERFACES) - new_count - mod_count

    print(f"Saved: {out}")
    print(f"  7 subsystems (SS1-SS7)")
    print(f"  {len(INTERFACES)} internal interfaces + {len(EXTERNALS)} external = {len(INTERFACES) + len(EXTERNALS)} total")
    print(f"  New: {new_count} | Modified: {mod_count} | Unchanged: {unchanged}")
    print(f"  Sheets: N2 Chart, Analysis, Legend")
