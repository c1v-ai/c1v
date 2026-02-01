"""Generate c1v-Identity UCBD Excel workbook following Cornell CESYS522 format."""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def style_header_row(ws, row, max_col=4):
    """Apply header styling to a UCBD column header row."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_section_label(ws, row, label, max_col=4):
    """Apply section label styling (Initial/Ending Conditions, Notes)."""
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    section_font = Font(name="Calibri", size=11, bold=True)
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill = section_fill
    cell.font = section_font
    for col in range(2, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = section_fill


def style_uc_title(ws, row, title):
    """Apply use case title styling."""
    title_font = Font(name="Calibri", size=13, bold=True, color="1F4E79")
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = title_font
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def apply_borders(ws, min_row, max_row, max_col=4):
    """Apply thin borders to a range of cells."""
    thin = Side(style="thin", color="B4C6E7")
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def write_ucbd(ws, start_row, uc_name, initial_conditions, steps, ending_conditions, notes, headers=None):
    """
    Write a single UCBD block starting at start_row.
    steps = list of tuples: (primary_actor_text, system_text, other_actors_text)
    Returns the next available row.
    """
    if headers is None:
        headers = ["Primary Actor", "The System", "Other Actors /\nInteraction Elements", "[Additional Actors]"]

    body_font = Font(name="Calibri", size=11)
    system_font = Font(name="Calibri", size=11, italic=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    row = start_row

    # Use Case Title
    style_uc_title(ws, row, f"Use Case Name: {uc_name}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    # Blank spacer
    row += 1

    # Initial Conditions
    style_section_label(ws, row, "Initial Conditions ", max_col=4)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    cond_cell = ws.cell(row=row, column=1, value=initial_conditions)
    cond_cell.font = body_font
    cond_cell.alignment = wrap
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    # Blank spacer
    row += 1

    # Column Headers
    header_row = row
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    style_header_row(ws, row, max_col=len(headers))
    row += 1

    # Steps
    steps_start = row
    for primary, system, other in steps:
        c1 = ws.cell(row=row, column=1, value=primary if primary else None)
        c1.font = body_font
        c1.alignment = wrap

        c2 = ws.cell(row=row, column=2, value=system if system else None)
        c2.font = system_font
        c2.alignment = wrap

        c3 = ws.cell(row=row, column=3, value=other if other else None)
        c3.font = body_font
        c3.alignment = wrap

        row += 1

    steps_end = row - 1
    apply_borders(ws, header_row, steps_end, max_col=4)

    # Blank spacer
    row += 1

    # Ending Conditions
    style_section_label(ws, row, "Ending Conditions ", max_col=4)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    end_cell = ws.cell(row=row, column=1, value=ending_conditions)
    end_cell.font = body_font
    end_cell.alignment = wrap
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    # Blank spacer
    row += 1

    # Notes
    style_section_label(ws, row, "Notes ", max_col=4)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    notes_cell = ws.cell(row=row, column=1, value=notes)
    notes_cell.font = body_font
    notes_cell.alignment = wrap
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    # Two blank spacer rows before next UC
    row += 2
    return row


def create_guidelines_sheet(wb):
    """Create the Guidelines for Requirements sheet."""
    ws = wb.create_sheet("Guidelines for Requirements")
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 70

    header_font = Font(name="Calibri", size=13, bold=True, color="1F4E79")
    body_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", size=11, bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    ws.cell(row=1, column=1, value="Properties and Guidelines for Writing Requirements").font = header_font

    guidelines = [
        ("1. Write as Shall Statements", 'meaning that the word shall is used in the requirement\n(i.e. "The system shall\u2026")'),
        ("2. Write as Correct Statements", "what you\u2019re saying is accurate"),
        ("3. Write Clear & Precise Statements", 'Share only one idea per requirement.\nIf you have the word "and" or similar conjunction in your requirements, it\u2019s considered better to split the requirement into two.'),
        ("4. Write Unambiguous Statements", "Let there only be one way to interpret the statement."),
        ("5. Write Objective Statements", "If any part of the requirement is subjective, qualify it with some quantifiable measure"),
        ("6. Write Verifiable Statements", "There is some measurable way you could say this requirement is met."),
        ("7. Write Consistent Statements", "This requirement does not contradict another requirement."),
        ("", ""),
        ("Additional guidelines:", "Some organizations may also include the following in their lists of best practices."),
        ("8. Implementation Independent", "functional, not structural"),
        ("9. Achievable", "feasible, meaning that it can be achieved"),
        ("10. Conforming", "consistent with regulations imposed by stakeholders and any other governing entity"),
    ]

    row = 4
    for prop, desc in guidelines:
        if prop:
            ws.cell(row=row, column=1, value=prop).font = bold_font
        ws.cell(row=row, column=2, value=desc).font = body_font
        ws.cell(row=row, column=1).alignment = wrap
        ws.cell(row=row, column=2).alignment = wrap
        row += 2


def main():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: UCBD ----
    ws = wb.active
    ws.title = "UCBD"

    # Column widths
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 30

    # Title block
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    sub_font = Font(name="Calibri", size=12, italic=True, color="4472C4")
    ws.cell(row=1, column=1, value="Use Case Behavioral Diagram").font = title_font
    ws.cell(row=2, column=1, value="c1v-Identity \u2014 AI-Native Customer Data Platform").font = sub_font
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")

    row = 4

    # ======================================================================
    # UC-01: Register & Connect Data Source
    # ======================================================================
    row = write_ucbd(
        ws, row,
        uc_name="UC-01 \u2014 Register & Connect Data Source",
        initial_conditions=(
            "1. The Data Administrator has an active c1v-Identity project.\n"
            "2. The Data Administrator has access credentials for the data source (CSV, database, or API).\n"
            "3. The System has fewer than the maximum allowed data sources for the project\u2019s plan tier."
        ),
        steps=[
            (
                'The Data Administrator selects "Add Data Source" from the Connectors page.',
                None,
                None,
            ),
            (
                None,
                "The System shall present available connector types (CSV Upload, Database, API) and prompt the user to select one.",
                "Connector Registry \u2014 Available connector types are loaded.",
            ),
            (
                "The Data Administrator selects a connector type and provides connection details (file upload, connection string, or API endpoint + credentials).",
                None,
                None,
            ),
            (
                None,
                "The System shall validate the connection details and attempt a test connection to the data source.",
                "External Data Source \u2014 Connection is tested.",
            ),
            (
                None,
                "The System shall retrieve and display the source schema (field names, data types, sample rows).",
                "Schema Inspector \u2014 Schema is extracted and displayed.",
            ),
            (
                "The Data Administrator reviews the schema and maps source fields to the canonical identity schema (email, first_name, last_name, phone, address, etc.).",
                None,
                None,
            ),
            (
                None,
                "The System shall store the field mapping configuration and associate it with the data source record.",
                "Configuration Store \u2014 Mapping is persisted.",
            ),
            (
                None,
                "The System shall assign a unique source identifier and register the data source in the project\u2019s source catalog.",
                "Source Catalog \u2014 New source entry is created.",
            ),
            (
                None,
                "The System shall display PII encryption simulation with visual feedback for sensitive fields detected in the schema.",
                "PII Detector \u2014 Sensitive fields are flagged.",
            ),
            (
                "The Data Administrator confirms the source registration.",
                None,
                None,
            ),
            (
                None,
                'The System shall update the Connectors dashboard to reflect the new source and enable the "Run Golden Records" button if two or more sources are registered.',
                "Dashboard \u2014 UI state is updated.",
            ),
        ],
        ending_conditions=(
            "1. A new data source is registered in the project\u2019s source catalog with a unique identifier.\n"
            "2. The field mapping from source schema to canonical identity schema is stored.\n"
            "3. The Connectors dashboard reflects the new source.\n"
            "4. If two or more sources exist, the \u201cRun Golden Records\u201d action is enabled."
        ),
        notes=(
            "1. Connection credentials are stored encrypted and never displayed in plain text after initial entry.\n"
            "2. The canonical identity schema includes at minimum: email, first_name, last_name, phone, postal_code, address.\n"
            "3. PII encryption simulation is for visual feedback in the MVP; production PII handling uses SHA256 hashing.\n"
            "4. CSV uploads are stored in a staging area; database/API sources are queried on demand."
        ),
    )

    # ======================================================================
    # UC-02: Run Data Quality Validation
    # ======================================================================
    row = write_ucbd(
        ws, row,
        uc_name="UC-02 \u2014 Run Data Quality Validation (DQ Gate)",
        initial_conditions=(
            "1. At least one data source is registered (UC-01 complete).\n"
            "2. A DQ contract exists for the source (auto-generated or manually configured).\n"
            "3. The Data Administrator has selected a source for validation."
        ),
        steps=[
            (
                'The Data Administrator selects a data source and clicks "Run DQ Validation".',
                None,
                None,
            ),
            (
                None,
                "The System shall load the DQ contract for the selected source, including schema rules, quality thresholds, and severity mappings.",
                "Contract Store \u2014 DQ contract YAML is loaded.",
            ),
            (
                None,
                "The System shall validate the source data against schema rules (required fields, data type checks).",
                "Schema Validator \u2014 Type and presence checks are executed.",
            ),
            (
                None,
                "The System shall execute quality checks on each field (null percentage thresholds, format validation for email/phone/postal patterns).",
                "Quality Rules Engine \u2014 Field-level quality metrics are computed.",
            ),
            (
                None,
                "The System shall verify PII consent scope for fields classified as personally identifiable.",
                "Consent Registry \u2014 Consent coverage is checked.",
            ),
            (
                None,
                "The System shall classify each check result by severity: GREEN (pass), AMBER (warning), or RED (critical).",
                "Severity Classifier \u2014 Results are color-coded.",
            ),
            (
                None,
                "The System shall apply environment-specific policy actions (PASS, WARN, or BLOCK) based on the gate policy configuration.",
                "Gate Policy \u2014 Environment rules (staging/prod) are applied.",
            ),
            (
                None,
                "The System shall generate a DQ validation report with per-field results, overall severity counts, and recommended actions.",
                "Report Generator \u2014 DQ report is produced.",
            ),
            (
                None,
                "The System shall log the validation event to the append-only DQ event audit trail with timestamp, source, environment, and outcome.",
                "Audit Trail \u2014 Event is appended to dq_events log.",
            ),
            (
                None,
                "The System shall trigger alerting stubs (Slack notification, JIRA ticket) for any non-GREEN results.",
                "Alert Router \u2014 Notifications are dispatched.",
            ),
            (
                "The Data Administrator reviews the DQ report and decides whether to proceed with identity resolution or remediate issues.",
                None,
                None,
            ),
        ],
        ending_conditions=(
            "1. A DQ validation report exists for the selected source with per-field severity ratings.\n"
            "2. The validation event is logged in the audit trail.\n"
            "3. If any RED results exist and the environment policy is BLOCK, the source is marked as \u201cBlocked\u201d and cannot proceed to identity resolution until remediated.\n"
            "4. Alerting stubs have been triggered for non-GREEN results."
        ),
        notes=(
            "1. DQ contracts can be auto-scaffolded from CSV schemas using the contract_scaffold module.\n"
            "2. Environment policies differ: staging may WARN on AMBER, while prod may BLOCK.\n"
            "3. PII type inference is rule-based (email regex, phone patterns, postal code formats).\n"
            "4. Industry presets (automotive, ecommerce, telehealth, generic) provide default contract templates."
        ),
    )

    # ======================================================================
    # UC-03: Resolve Identities (Identity Unify)
    # ======================================================================
    row = write_ucbd(
        ws, row,
        uc_name="UC-03 \u2014 Resolve Identities (Identity Unify)",
        initial_conditions=(
            "1. Two or more data sources are registered and have passed DQ validation (UC-01, UC-02 complete).\n"
            "2. The Unify Policy configuration is loaded (match thresholds, blocking strategies, scoring weights).\n"
            "3. The Data Administrator has triggered \u201cRun Golden Records\u201d."
        ),
        steps=[
            (
                'The Data Administrator clicks "Run Golden Records" from the Dashboard.',
                None,
                None,
            ),
            (
                None,
                "The System shall load all registered source datasets and apply multi-source canonicalization to map each source\u2019s fields to the common canonical schema.",
                "Source Catalog \u2014 Source data is loaded and transformed.",
            ),
            (
                None,
                "The System shall normalize identity fields: email (Gmail dot removal, plus addressing), phone (digit extraction, country code handling), name (case normalization), address (basic standardization).",
                "Normalization Engine \u2014 Fields are normalized per configured rules.",
            ),
            (
                None,
                "The System shall generate blocking keys for each record using configured blocking strategies (email domain + last 4 chars, phone last 7 digits, name + FSA, exact email match).",
                "Blocking Engine \u2014 Blocking keys are computed and indexed.",
            ),
            (
                None,
                "The System shall compare records within each blocking group using weighted similarity scoring (email: 0.9, phone: 0.7, name: 0.5, postal: 0.2).",
                "Scoring Engine \u2014 Pairwise similarity scores are computed.",
            ),
            (
                None,
                "The System shall classify each pair by match decision: auto_merge (\u22650.9), needs_review (\u22650.7), or no_match (<0.7).",
                "Decision Classifier \u2014 Match decisions are assigned.",
            ),
            (
                None,
                "The System shall cluster auto_merge pairs into identity groups using transitive closure.",
                "Clustering Engine \u2014 Identity clusters are formed.",
            ),
            (
                None,
                "The System shall apply survivorship rules to each cluster to produce a golden record (selecting the most complete/recent value for each field).",
                "Survivorship Rules \u2014 Golden records are constructed.",
            ),
            (
                None,
                "The System shall assign a unique golden record identifier (UID) to each resolved identity.",
                "UID Generator \u2014 Persistent UIDs are assigned.",
            ),
            (
                None,
                "The System shall export golden records to the golden contacts output and log all match decisions to the unify events audit trail.",
                "Output Store \u2014 golden_contacts.csv and unify_events.csv are written.",
            ),
            (
                None,
                "The System shall compute and display unification metrics: duplicate rate, auto-merge rate, compression ratio, and needs-review count.",
                "Metrics Dashboard \u2014 Unification KPIs are displayed.",
            ),
            (
                None,
                "The System shall flag pairs classified as needs_review for human review (UC-04).",
                "Review Queue \u2014 Uncertain matches are queued.",
            ),
        ],
        ending_conditions=(
            "1. Golden records are produced with unique UIDs for all auto_merge clusters.\n"
            "2. All match decisions (auto_merge, needs_review, no_match) are logged in the unify events audit trail.\n"
            "3. Unification metrics are computed and displayed on the Dashboard.\n"
            "4. Records classified as needs_review are queued for human review (UC-04)."
        ),
        notes=(
            "1. Blocking strategies reduce the comparison space from O(n\u00b2) to manageable candidate pairs.\n"
            "2. Scoring weights are configurable via unify_policy.yaml.\n"
            "3. Survivorship rules can be source-priority-based or recency-based, depending on configuration.\n"
            "4. PII masking is applied to golden record output based on RBAC role of the requesting user.\n"
            "5. The compression ratio = total source records / golden records; higher is better."
        ),
    )

    # ======================================================================
    # UC-04: Review & Resolve Merge Conflicts
    # ======================================================================
    row = write_ucbd(
        ws, row,
        uc_name="UC-04 \u2014 Review & Resolve Merge Conflicts",
        initial_conditions=(
            "1. Identity resolution has been run (UC-03 complete).\n"
            "2. One or more record pairs are classified as needs_review (score \u22650.7 and <0.9).\n"
            "3. The Data Administrator or CX Lead navigates to the Merge Errors tab."
        ),
        steps=[
            (
                'The reviewer opens the "Record Merge Errors" tab on the Dashboard.',
                None,
                None,
            ),
            (
                None,
                "The System shall display an error summary showing counts by severity level (RED, YELLOW, GREEN) and total needs-review pairs.",
                "Error Summary Dashboard \u2014 Severity breakdown is shown.",
            ),
            (
                None,
                "The System shall present the error breakdown table with error type, count, and details for each severity level.",
                "Error Table \u2014 Categorized errors are listed.",
            ),
            (
                "The reviewer filters errors by severity (RED/YELLOW/GREEN) to prioritize resolution.",
                None,
                None,
            ),
            (
                None,
                "The System shall filter and re-display the error list based on the selected severity.",
                None,
            ),
            (
                'The reviewer selects a needs-review pair and clicks "Resolve".',
                None,
                None,
            ),
            (
                None,
                "The System shall display both records side-by-side with field-level comparison, similarity scores, and the blocking key that linked them.",
                "Record Comparison View \u2014 Side-by-side records are shown.",
            ),
            (
                "The reviewer selects the correct action: Merge (confirm match), Split (reject match), or Edit (manually adjust golden record fields).",
                None,
                None,
            ),
            (
                None,
                "The System shall apply the reviewer\u2019s decision: merge the records into a single golden record, keep them separate, or update fields as specified.",
                "Resolution Engine \u2014 Decision is applied to the identity graph.",
            ),
            (
                None,
                "The System shall log the resolution decision with the reviewer\u2019s identity, timestamp, action taken, and rationale.",
                "Audit Trail \u2014 Resolution event is logged.",
            ),
            (
                None,
                "The System shall update the Dashboard metrics (needs-review count, duplicate rate, compression ratio) to reflect the resolution.",
                "Metrics Dashboard \u2014 KPIs are recalculated.",
            ),
        ],
        ending_conditions=(
            "1. The selected needs-review pair has been resolved (merged, split, or edited).\n"
            "2. The resolution decision is logged in the audit trail.\n"
            "3. Dashboard metrics are updated to reflect the change.\n"
            "4. The needs-review count has decreased by one."
        ),
        notes=(
            '1. "Solve Record Merge Errors" CTA button on the Golden Records dashboard links directly to this workflow.\n'
            "2. Merge conflicts may arise from ambiguous data (e.g., same name but different phone numbers across sources).\n"
            "3. Split decisions create a permanent \u201cdo not merge\u201d rule for the pair to prevent future false positives.\n"
            "4. PII masking is applied based on the reviewer\u2019s RBAC role; some fields may be partially masked."
        ),
    )

    # ======================================================================
    # UC-05: Ask the Data (Natural Language Query)
    # ======================================================================
    row = write_ucbd(
        ws, row,
        uc_name="UC-05 \u2014 Ask the Data (Natural Language Query)",
        initial_conditions=(
            "1. Golden records exist (UC-03 complete).\n"
            "2. The user has an active project with at least one completed identity resolution run.\n"
            "3. The user navigates to the \u201cAsk the Data\u201d tab."
        ),
        steps=[
            (
                "The user types a natural-language question (e.g., \u201cRevenue by day for the last 30 days\u201d).",
                None,
                None,
            ),
            (
                None,
                "The System shall route the natural-language question through template matching to identify a known query pattern.",
                "Query Router \u2014 Templates are matched against the question.",
            ),
            (
                None,
                "The System shall generate a SQL query from the matched template, injecting appropriate parameters (date ranges, aggregations, filters).",
                "SQL Generator \u2014 Parameterized SQL is produced.",
            ),
            (
                None,
                "The System shall validate the generated SQL: SELECT-only, no wildcard (*), allowlisted views only, partition window present, LIMIT clause enforced.",
                "SQL Validator \u2014 Query safety checks are applied.",
            ),
            (
                None,
                "The System shall enforce cost guardrails: byte cap \u22641 GB (dry-run estimate), row cap \u226450,000 rows.",
                "Cost Guard \u2014 Resource limits are verified.",
            ),
            (
                None,
                "The System shall inject a default time window (90 days) if the user\u2019s question does not specify a date range.",
                "Default Injector \u2014 Missing parameters are defaulted.",
            ),
            (
                None,
                "The System shall execute the validated SQL query against the allowlisted views (dm.metrics_*, dm.dim_customer_masked, dm.fact_events_view).",
                "Query Executor \u2014 SQL is run against the data store.",
            ),
            (
                None,
                "The System shall render the query results as a dynamic chart and data table in the UI.",
                "Chart Renderer \u2014 Visualization is generated.",
            ),
            (
                None,
                "The System shall log the query (question text, generated SQL, execution time, row count) to the query log.",
                "Query Log \u2014 Usage is recorded.",
            ),
            (
                "The user reviews the results and optionally refines the question.",
                None,
                None,
            ),
        ],
        ending_conditions=(
            "1. The user\u2019s natural-language question has been answered with a chart and data table.\n"
            "2. The query is logged in the query log.\n"
            "3. All guardrails (SELECT-only, byte cap, row cap, allowlisted views) have been enforced."
        ),
        notes=(
            "1. Template coverage target is \u226570% of common analytical questions.\n"
            "2. If no template matches, the System may use OpenAI integration to generate SQL (Dashboard Builder feature).\n"
            "3. Allowlisted views are configured in ask_data/constants.py.\n"
            "4. In simulation mode (ASKDATA_SIMULATE=1), mock data is returned instead of querying a real data store."
        ),
    )

    # ======================================================================
    # UC-06: Match Identity via API
    # ======================================================================
    row = write_ucbd(
        ws, row,
        uc_name="UC-06 \u2014 Match Identity via API (Real-Time)",
        initial_conditions=(
            "1. The c1v-Identity API server is running.\n"
            "2. Golden records exist from at least one identity resolution run.\n"
            "3. An external system (AI agent, chatbot, CRM) sends a POST request to the /match endpoint."
        ),
        steps=[
            (
                "The external system sends a POST /match request with two records (record1, record2) containing identity fields (email, first_name, last_name, phone).",
                None,
                "External System \u2014 AI Agent, Chatbot, or CRM sends the match request.",
            ),
            (
                None,
                "The System shall validate the incoming request payload for required fields and correct data types.",
                "Request Validator \u2014 Input schema is enforced.",
            ),
            (
                None,
                "The System shall normalize identity fields on both records using the same normalization rules as batch resolution (email, phone, name normalization).",
                "Normalization Engine \u2014 Fields are normalized.",
            ),
            (
                None,
                "The System shall compute pairwise similarity scores using weighted scoring (email: 0.9, phone: 0.7, name: 0.5, postal: 0.2).",
                "Scoring Engine \u2014 Similarity score is computed.",
            ),
            (
                None,
                "The System shall determine match result (match: true/false) and confidence score based on configured thresholds.",
                "Decision Classifier \u2014 Match decision is made.",
            ),
            (
                None,
                "The System shall determine the primary match reason (e.g., exact-email-match, fuzzy-name-phone-match, no-match).",
                "Reason Classifier \u2014 Match reason is identified.",
            ),
            (
                None,
                "The System shall return a JSON response containing: match (boolean), confidence (float 0\u20131), and reason (string).",
                "API Response \u2014 Result is returned to the caller.",
            ),
            (
                None,
                "The System shall log the API call (timestamp, request hash, confidence, decision) for monitoring and audit.",
                "API Audit Log \u2014 Request is recorded.",
            ),
        ],
        ending_conditions=(
            "1. The external system has received a JSON response with match, confidence, and reason fields.\n"
            "2. The API call is logged for audit and monitoring.\n"
            "3. The API response latency is within acceptable bounds for real-time use."
        ),
        notes=(
            "1. This is the real-time equivalent of the batch identity resolution (UC-03) \u2014 same normalization and scoring logic.\n"
            "2. The API is designed for AI agent integration: LangChain IdentityTool and MCP Server both call this endpoint.\n"
            "3. Scoring weights and thresholds match the batch unify_policy.yaml configuration.\n"
            "4. Rate limiting and authentication are handled at the API gateway layer.\n"
            "5. Response format: {\"match\": true, \"confidence\": 0.95, \"reason\": \"exact-email-match\"}"
        ),
    )

    # ======================================================================
    # UC-07: Launch Privacy-Safe Campaign
    # ======================================================================
    row = write_ucbd(
        ws, row,
        uc_name="UC-07 \u2014 Launch Privacy-Safe Campaign",
        initial_conditions=(
            "1. Golden records exist with consent metadata (UC-03 complete).\n"
            "2. The Campaign Manager has defined campaign parameters (audience, channel, timing).\n"
            "3. Privacy policies and consent rules are configured in the System."
        ),
        steps=[
            (
                "The Campaign Manager selects audience criteria (segment, geography, purchase history) for the campaign.",
                None,
                None,
            ),
            (
                None,
                "The System shall query the golden records using the audience criteria and return matching identity profiles.",
                "Golden Record Store \u2014 Audience segment is extracted.",
            ),
            (
                None,
                "The System shall filter the audience to include only records with valid consent for the specified channel and purpose.",
                "Consent Registry \u2014 Consent validation is applied.",
            ),
            (
                None,
                "The System shall tokenize personally identifiable fields before export, replacing PII with opaque tokens.",
                "Tokenization Engine \u2014 PII is replaced with tokens.",
            ),
            (
                None,
                "The System shall enforce export caps (maximum audience size per channel per time window) to prevent over-communication.",
                "Export Cap Policy \u2014 Volume limits are enforced.",
            ),
            (
                None,
                "The System shall check brand-safety rules and apply crisis throttle if an active brand-safety freeze is in effect.",
                "Brand Safety Gate \u2014 Freeze status is checked.",
            ),
            (
                "The Campaign Manager reviews the tokenized audience list and approves the campaign for activation.",
                None,
                None,
            ),
            (
                None,
                "The System shall route the tokenized audience to the activation destination (email platform, ad network, etc.) with health checks.",
                "Activation Destination \u2014 Audience is delivered.",
            ),
            (
                None,
                "The System shall log the campaign activation event with audience size, consent coverage, and destination details.",
                "Campaign Audit Log \u2014 Activation event is recorded.",
            ),
        ],
        ending_conditions=(
            "1. A privacy-safe, tokenized audience has been delivered to the activation destination.\n"
            "2. All records in the audience have valid consent for the channel and purpose.\n"
            "3. Export caps and brand-safety rules have been enforced.\n"
            "4. The campaign activation is logged in the audit trail."
        ),
        notes=(
            "1. Tokenization ensures that downstream systems never receive raw PII.\n"
            "2. Consent enforcement follows the privacy-first edge architecture (CPO, Security/IAM, Data Steward oversight).\n"
            "3. Crisis throttle (EXT-CRISIS) can halt all campaign activations when triggered by authorized personnel.\n"
            "4. This use case includes INC-CONSENT, INC-TOKEN, INC-ALERT, and INC-SLA from the context model."
        ),
    )

    # ======================================================================
    # UC-08: Process DSAR / Privacy Request
    # ======================================================================
    row = write_ucbd(
        ws, row,
        uc_name="UC-08 \u2014 Process DSAR / Privacy Request",
        initial_conditions=(
            "1. A Data Subject Access Request (DSAR) has been received (via email, web form, or legal channel).\n"
            "2. The Data Steward or CPO has authenticated the request and confirmed the data subject\u2019s identity.\n"
            "3. The System has golden records that may contain the data subject\u2019s information."
        ),
        steps=[
            (
                "The Data Steward enters the DSAR details (subject identity, request type: access/deletion/portability, and deadline).",
                None,
                None,
            ),
            (
                None,
                "The System shall search the golden records and all source records for any data matching the data subject\u2019s identity (email, phone, name).",
                "Identity Search \u2014 All linked records are discovered.",
            ),
            (
                None,
                "The System shall catalog all discovered records by source, field, and data category (PII, behavioral, transactional).",
                "Data Catalog \u2014 Record inventory is compiled.",
            ),
            (
                None,
                "The System shall apply region-specific residency rules to determine which data is subject to the request jurisdiction.",
                "Residency Rules \u2014 Jurisdictional scope is applied.",
            ),
            (
                None,
                "The System shall generate a data subject report listing all stored data, its sources, purposes, and retention periods.",
                "DSAR Report Generator \u2014 Subject report is produced.",
            ),
            (
                "The Data Steward reviews the report and confirms the action (export data, delete data, or restrict processing).",
                None,
                None,
            ),
            (
                None,
                "The System shall execute the confirmed action: export a portable data package, mark records for deletion, or flag records as restricted.",
                "Action Executor \u2014 DSAR action is applied.",
            ),
            (
                None,
                "The System shall log the DSAR processing decision with the authorizer\u2019s identity, timestamp, scope, and action taken.",
                "Decision Log \u2014 DSAR decision is recorded (INC-DECISION).",
            ),
            (
                None,
                "The System shall update the golden record graph to reflect any deletions or restrictions.",
                "Identity Graph \u2014 Graph is updated to reflect changes.",
            ),
        ],
        ending_conditions=(
            "1. The DSAR has been fully processed within the required deadline.\n"
            "2. The data subject report has been generated and delivered (for access requests).\n"
            "3. Records have been deleted or restricted as requested (for deletion/restriction requests).\n"
            "4. The DSAR decision is logged in the audit trail with full traceability."
        ),
        notes=(
            "1. DSAR processing includes INC-AUTH, INC-CATALOG, INC-TOKEN, and INC-DECISION from the context model.\n"
            "2. Residency routing (EXT-RESIDENCY) ensures data handling complies with regional regulations (GDPR, CCPA, PIPEDA).\n"
            "3. Deletion requests may require dual-control approval (INC-APPROVE) for high-risk data categories.\n"
            "4. The System must complete DSAR processing within the regulatory deadline (typically 30 days for GDPR)."
        ),
    )

    # ---- Sheet 2: Guidelines ----
    create_guidelines_sheet(wb)

    # ---- Save ----
    output_path = "/Users/davidancor/coding/c1v/apps/c1v-identity/c1v-identity-UCBD.xlsx"
    wb.save(output_path)
    print(f"UCBD Excel saved to: {output_path}")


if __name__ == "__main__":
    main()
