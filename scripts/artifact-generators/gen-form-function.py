#!/usr/bin/env python3
"""gen-form-function.py — Crawley Form-Function mapping generator.

Renders ``form_function_map.v1.json`` to:
  - xlsx : function × form quality matrix (Pugh-style, via openpyxl)
  - svg  : bipartite graph function↔form (networkx + matplotlib)
  - mmd  : Mermaid concept expansion tree per Crawley (Form branch, Function branch)

Schema: permissive stub ``form-function-map.stub.schema.json`` until T5 lands
the canonical schema.

Instance shape (loose):
{
  "functions": [{"id": "F1", "verb": "store", "object": "orders",
                 "parentId": null}],
  "forms":     [{"id": "M1", "name": "PostgreSQL", "parentId": null}],
  "mappings":  [{"functionId": "F1", "formId": "M1", "quality": 5}],   # 1-5 Likert
  "conceptExpansion": {
    "function": {"id": "F0", "label": "persist data",
                 "children": [{"id": "F1", "label": "store orders"}]},
    "form":     {"id": "M0", "label": "database",
                 "children": [{"id": "M1", "label": "PostgreSQL"}]}
  }
}
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from common import extender_init  # noqa: F401,E402
from common.runner import run_generator  # noqa: E402

import matplotlib  # noqa: E402
matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
import networkx as nx  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402


_FONT = {"family": "DejaVu Sans"}


def _save_svg(fig, path: Path) -> None:
    fig.savefig(path, format="svg", bbox_inches="tight")
    plt.close(fig)


def _xlsx_matrix(instance: dict, out_path: Path, warnings: list[str]) -> bool:
    functions = instance.get("functions") or []
    forms = instance.get("forms") or []
    mappings = instance.get("mappings") or []
    if not functions or not forms:
        warnings.append(
            "gen-form-function: xlsx skipped — missing functions or forms"
        )
        return False
    lookup: dict[tuple[str, str], Any] = {}
    for m in mappings:
        lookup[(m.get("functionId"), m.get("formId"))] = m.get("quality")

    wb = Workbook()
    ws = wb.active
    ws.title = "Function × Form"
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="DDDDDD")

    ws.cell(row=1, column=1, value="Function \\ Form").font = hdr_font
    ws.cell(row=1, column=1).fill = hdr_fill
    for j, form in enumerate(forms, start=2):
        cell = ws.cell(row=1, column=j, value=form.get("name", form.get("id")))
        cell.font = hdr_font
        cell.fill = hdr_fill
    for i, fn in enumerate(functions, start=2):
        fid = fn.get("id")
        label = fn.get("verb", "") + " " + fn.get("object", "")
        label = label.strip() or fid
        ws.cell(row=i, column=1, value=label).font = hdr_font
        for j, form in enumerate(forms, start=2):
            ws.cell(row=i, column=j, value=lookup.get((fid, form.get("id"))))
    for col_idx in range(1, len(forms) + 2):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
    wb.save(out_path)
    return True


def _svg_bipartite(instance: dict, out_path: Path, warnings: list[str]) -> bool:
    functions = instance.get("functions") or []
    forms = instance.get("forms") or []
    mappings = instance.get("mappings") or []
    if not functions or not forms:
        warnings.append(
            "gen-form-function: bipartite svg skipped — missing functions or forms"
        )
        return False
    g = nx.Graph()
    for fn in functions:
        fid = fn.get("id")
        label = (fn.get("verb", "") + " " + fn.get("object", "")).strip() or fid
        g.add_node(fid, side="function", label=label)
    for fm in forms:
        g.add_node(fm.get("id"), side="form", label=fm.get("name", fm.get("id")))
    for m in mappings:
        q = m.get("quality") or 0
        if q:
            g.add_edge(m.get("functionId"), m.get("formId"), quality=q)

    fn_nodes = [n for n, d in g.nodes(data=True) if d.get("side") == "function"]
    fm_nodes = [n for n, d in g.nodes(data=True) if d.get("side") == "form"]
    pos: dict[str, tuple[float, float]] = {}
    for i, n in enumerate(fn_nodes):
        pos[n] = (0.0, -i)
    for i, n in enumerate(fm_nodes):
        pos[n] = (1.5, -i)

    fig, ax = plt.subplots(figsize=(9, max(4, 0.55 * max(len(fn_nodes), len(fm_nodes)))))
    widths = [0.5 + 0.6 * g[u][v].get("quality", 1) for u, v in g.edges()]
    nx.draw_networkx_edges(g, pos, ax=ax, edge_color="#5998C5", width=widths, alpha=0.7)
    nx.draw_networkx_nodes(g, pos, nodelist=fn_nodes, node_color="#F18F01",
                           edgecolors="#0B2C29", node_size=1200, ax=ax)
    nx.draw_networkx_nodes(g, pos, nodelist=fm_nodes, node_color="#FBFCFC",
                           edgecolors="#0B2C29", node_size=1200, ax=ax)
    labels = {n: (d.get("label") or n)[:22] for n, d in g.nodes(data=True)}
    nx.draw_networkx_labels(
        g, pos, labels=labels, font_size=7, ax=ax, font_family="DejaVu Sans"
    )
    ax.set_title("Function ↔ Form (edge width = quality)", **_FONT)
    ax.set_axis_off()
    _save_svg(fig, out_path)
    return True


def _mmd_concept_tree(instance: dict, out_path: Path, warnings: list[str]) -> bool:
    expansion = instance.get("conceptExpansion") or {}
    function_tree = expansion.get("function")
    form_tree = expansion.get("form")
    if not function_tree and not form_tree:
        warnings.append(
            "gen-form-function: mmd skipped — conceptExpansion missing"
        )
        return False

    def _safe_id(node_id: str) -> str:
        return "".join(ch if ch.isalnum() else "_" for ch in str(node_id))

    lines: list[str] = ["graph LR"]

    def _walk(tree: dict, prefix: str) -> None:
        if not tree:
            return
        nid = f"{prefix}_{_safe_id(tree.get('id', 'root'))}"
        label = str(tree.get("label", tree.get("id", "?"))).replace('"', "'")
        lines.append(f'    {nid}["{label}"]')
        for child in tree.get("children") or []:
            cid = f"{prefix}_{_safe_id(child.get('id', 'c'))}"
            clabel = str(child.get("label", child.get("id", "?"))).replace('"', "'")
            lines.append(f'    {cid}["{clabel}"]')
            lines.append(f"    {nid} --> {cid}")
            _walk(child, prefix)

    if function_tree:
        lines.append("    subgraph Function_Branch")
        _walk(function_tree, "fn")
        lines.append("    end")
    if form_tree:
        lines.append("    subgraph Form_Branch")
        _walk(form_tree, "fm")
        lines.append("    end")

    with out_path.open("w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")
    return True


def render(instance, output_dir: Path, targets, options, warnings):
    base = options.get("outputBasename") or "form_function_map"
    generated: list[dict[str, str]] = []

    if "xlsx" in targets:
        out = output_dir / f"{base}.xlsx"
        if _xlsx_matrix(instance, out, warnings):
            generated.append({"target": "xlsx", "path": str(out)})

    if "svg" in targets:
        out = output_dir / f"{base}.bipartite.svg"
        try:
            if _svg_bipartite(instance, out, warnings):
                generated.append({"target": "svg", "path": str(out)})
        except Exception as exc:
            warnings.append(
                f"gen-form-function: bipartite svg failed: {type(exc).__name__}: {exc}"
            )

    if "mmd" in targets:
        out = output_dir / f"{base}.concept-tree.mmd"
        try:
            if _mmd_concept_tree(instance, out, warnings):
                generated.append({"target": "mmd", "path": str(out)})
        except Exception as exc:
            warnings.append(
                f"gen-form-function: mmd failed: {type(exc).__name__}: {exc}"
            )

    if not generated:
        warnings.append("gen-form-function: no outputs produced (check targets/instance)")
    return generated


if __name__ == "__main__":
    raise SystemExit(run_generator(generator_name="gen-form-function", render_fn=render))
