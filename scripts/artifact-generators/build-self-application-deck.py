#!/usr/bin/env python3
"""build-self-application-deck.py — single entry point that regenerates
every artifact in plans/dogfooding-artifact-deck/portfolio/.

Order of operations:
  1. Per-module flat tables (M1/M2/M3/M4/M5)         — gen-self-app-flat-tables.py
  2. Per-module data slides (M1/M3/M4/M5/M8)         — gen-self-app-data-slides.py
  3. M2 UCBD slide deck (David's data-driven gen)    — generate_ucbd_pptx.py
  4. Synthesis capstone (PPT + xlsx + 5 Mermaid)     — gen-self-app-synthesis.py
  5. Master PPT assembler                             — assemblers/assemble-master-pptx.py
  6. Master XLSX assembler                            — assemblers/assemble-master-xlsx.py
  7. Build manifest written to portfolio/build-manifest.json

Idempotent. Safe to re-run. Times each stage.

Usage:
    scripts/artifact-generators/.venv/bin/python \\
        scripts/artifact-generators/build-self-application-deck.py
    # or with custom output dir:
    ... build-self-application-deck.py --output-dir <path>

Prereq: mmdc on PATH (`npm install -g @mermaid-js/mermaid-cli`).
"""
from __future__ import annotations

import argparse
import datetime
import json
import os
import shutil
import subprocess
import sys
import time
from pathlib import Path
from typing import Any, Dict, List

REPO_ROOT = Path(__file__).resolve().parents[2]
HERE = Path(__file__).resolve().parent

DEFAULT_OUTPUT_DIR = REPO_ROOT / 'plans' / 'dogfooding-artifact-deck' / 'portfolio'

PYTHON = sys.executable

UCBD_SCRIPT = (
    REPO_ROOT
    / 'apps/product-helper/.planning/phases'
    / '14-artifact-publishing-json-excel-ppt-pdf'
    / '2-dev-sys-reqs-for-kb-llm-software'
    / 'generate_ucbd_pptx.py'
)
M2_PROJECT = REPO_ROOT / 'system-design/kb-upgrade-v2/module-2-requirements'
SYNTHESIS_INPUT = (
    REPO_ROOT
    / '.planning/runs/self-application/synthesis/architecture_recommendation.v1.json'
)


def _check_prereqs() -> List[str]:
    errors: List[str] = []
    if not shutil.which('mmdc'):
        errors.append(
            "mmdc not found on PATH — install with `npm install -g @mermaid-js/mermaid-cli`"
        )
    for path in (UCBD_SCRIPT, M2_PROJECT, SYNTHESIS_INPUT):
        if not Path(path).exists():
            errors.append(f"required input missing: {path}")
    return errors


def _run(stage: str, cmd: List[str], log: List[Dict[str, Any]]) -> None:
    t0 = time.monotonic()
    print(f'[{stage}] {" ".join(str(c) for c in cmd[:6])} ...', flush=True)
    result = subprocess.run(cmd, capture_output=True, text=True, check=False)
    elapsed = time.monotonic() - t0
    entry = {
        'stage': stage,
        'cmd': cmd,
        'returncode': result.returncode,
        'elapsed_s': round(elapsed, 2),
    }
    if result.returncode != 0:
        entry['stderr'] = result.stderr[-2000:]
        entry['stdout_tail'] = result.stdout[-500:]
        log.append(entry)
        print(f'[{stage}] FAILED in {elapsed:.1f}s', file=sys.stderr)
        print(result.stderr, file=sys.stderr)
        raise SystemExit(result.returncode)
    log.append(entry)
    print(f'[{stage}] ok ({elapsed:.1f}s)')


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument('--output-dir', default=str(DEFAULT_OUTPUT_DIR))
    ap.add_argument('--skip-mermaid', action='store_true',
                    help='skip mmdc check + synthesis Mermaid rasterization')
    args = ap.parse_args()

    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    if not args.skip_mermaid:
        prereq_errors = _check_prereqs()
        if prereq_errors:
            for e in prereq_errors:
                print(f'PREREQ: {e}', file=sys.stderr)
            raise SystemExit(2)

    started_at = datetime.datetime.now(datetime.timezone.utc).isoformat()
    log: List[Dict[str, Any]] = []
    t_total = time.monotonic()

    # Run from repo root so relative source paths resolve.
    os.chdir(REPO_ROOT)

    _run(
        'flat-tables',
        [PYTHON, str(HERE / 'gen-self-app-flat-tables.py'),
         '--output-dir', str(output_dir)],
        log,
    )
    _run(
        'data-slides',
        [PYTHON, str(HERE / 'gen-self-app-data-slides.py'),
         '--output-dir', str(output_dir)],
        log,
    )
    _run(
        'm2-ucbd-pptx',
        [PYTHON, str(UCBD_SCRIPT),
         '--project', str(M2_PROJECT),
         '--output', str(output_dir / 'M2-UCBDs.pptx')],
        log,
    )
    _run(
        'synthesis',
        [PYTHON, str(HERE / 'gen-self-app-synthesis.py'),
         '--input', str(SYNTHESIS_INPUT),
         '--output-dir', str(output_dir)],
        log,
    )
    _run(
        'assemble-master-pptx',
        [PYTHON, str(HERE / 'assemblers' / 'assemble-master-pptx.py'),
         '--manifest', str(HERE / 'assemblers' / 'master-deck.manifest.json'),
         '--output', str(output_dir / 'c1v-self-application.pptx')],
        log,
    )
    _run(
        'assemble-master-xlsx',
        [PYTHON, str(HERE / 'assemblers' / 'assemble-master-xlsx.py'),
         '--manifest', str(HERE / 'assemblers' / 'master-workbook.manifest.json'),
         '--output', str(output_dir / 'c1v-self-application.xlsx')],
        log,
    )

    total_elapsed = time.monotonic() - t_total

    # Summary stats
    master_pptx = output_dir / 'c1v-self-application.pptx'
    master_xlsx = output_dir / 'c1v-self-application.xlsx'

    pptx_slides = None
    xlsx_sheets = None
    try:
        from pptx import Presentation
        from openpyxl import load_workbook
        pptx_slides = len(Presentation(str(master_pptx)).slides)
        xlsx_sheets = len(load_workbook(str(master_xlsx)).sheetnames)
    except Exception:
        pass

    manifest = {
        'started_at': started_at,
        'finished_at': datetime.datetime.now(datetime.timezone.utc).isoformat(),
        'total_elapsed_s': round(total_elapsed, 2),
        'output_dir': str(output_dir),
        'master_pptx': {
            'path': str(master_pptx),
            'size_bytes': master_pptx.stat().st_size if master_pptx.exists() else None,
            'slide_count': pptx_slides,
        },
        'master_xlsx': {
            'path': str(master_xlsx),
            'size_bytes': master_xlsx.stat().st_size if master_xlsx.exists() else None,
            'sheet_count': xlsx_sheets,
        },
        'stages': log,
    }

    manifest_path = output_dir / 'build-manifest.json'
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding='utf-8')

    print()
    print(f'BUILD COMPLETE in {total_elapsed:.1f}s')
    print(f'  master pptx: {master_pptx} ({pptx_slides or "?"} slides)')
    print(f'  master xlsx: {master_xlsx} ({xlsx_sheets or "?"} sheets)')
    print(f'  manifest:    {manifest_path}')


if __name__ == '__main__':
    main()
