#!/usr/bin/env python3
"""gen-qfd.py — House of Quality (QFD) xlsx generator.

Migrated from:
    system-design/kb-upgrade-v2/module-6-qfd/write_xlsx.py

Changes vs legacy:
  - Drops AppleScript dependency (legacy used osascript + Microsoft Excel).
    Replaced with pure openpyxl — cross-platform, CI-friendly.
  - Conforms to v2 §15.3 input/output contract via common/runner.py.
  - Reads cell writes from ``instanceJson`` (legacy read a sibling c1v_QFD.json).

Target: xlsx.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from common.runner import run_generator  # noqa: E402

from openpyxl import load_workbook  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402


def _ec_col(n: int) -> str:
    return get_column_letter(n + 4)


def _plan_writes(data: dict) -> list[tuple[str, object]]:
    writes: list[tuple[str, object]] = []

    md = data["metadata"]
    writes += [
        ("B2", md["project_title"]),
        ("B4", md["developed_by"]),
        ("B6", md["last_updated"]),
    ]

    # Front Porch rows 33-38
    for i, pc in enumerate(data["front_porch"]):
        row = 33 + i
        writes += [
            (f"B{row}", pc["full_attribute"]),
            (f"C{row}", pc["short_name"]),
            (f"D{row}", pc["relative_importance"]),
        ]

    # Second Floor
    for ec in data["second_floor"]:
        n = ec["ec_id"]
        col = _ec_col(n)
        writes += [
            (f"{col}30", ec["name"]),
            (f"{col}31", ec["direction_of_change"]),
        ]

    # Main Floor
    pc_order = ["PC.1", "PC.2", "PC.3", "PC.4", "PC.5", "PC.6"]
    for pc_i, pc_key in enumerate(pc_order):
        row = 33 + pc_i
        for n in range(1, 19):
            val = data["main_floor"][pc_key].get(f"EC{n}", 0)
            if val != 0:
                writes.append((f"{_ec_col(n)}{row}", val))

    # Roof (lower triangle)
    for key, v in data.get("roof", {}).items():
        if key.startswith("_"):
            continue
        i, j = map(int, key.replace("EC", "").split("_"))
        assert i < j, f"roof key {key!r}: expected i<j"
        cell = f"{get_column_letter(i + 4)}{j + 2}"
        writes.append((cell, v))

    # Back Porch
    for pc_i, pc_key in enumerate(pc_order):
        row = 33 + pc_i
        bp = data["back_porch"][pc_key]
        writes += [
            (f"AE{row}", bp["A_low"]),
            (f"AF{row}", bp["A_high"]),
            (f"AG{row}", bp["A_target"]),
            (f"AH{row}", bp["B"]),
            (f"AI{row}", bp["C"]),
        ]

    # Basement
    writes.append(("C46", data["back_porch"]["_competitor_b_name"]))
    writes.append(("C47", data["back_porch"]["_competitor_c_name"]))
    for n in range(1, 19):
        col = _ec_col(n)
        b = data["basement"][f"EC{n}"]
        writes += [
            (f"{col}45", b["unit"]),
            (f"{col}46", b["competitor_b"]),
            (f"{col}47", b["competitor_c"]),
            (f"{col}48", b["external_threshold"]),
            (f"{col}49", b["target"]),
            (f"{col}53", b["technical_difficulty"]),
            (f"{col}54", b["estimated_cost"]),
        ]

    return writes


def render(instance, output_dir: Path, targets, options, warnings):
    if "xlsx" not in targets:
        warnings.append("gen-qfd: no xlsx target requested; nothing to do")
        return []

    template = options.get("templatePath")
    if not template:
        # Default template lives next to the legacy script.
        default_template = (
            Path(__file__).resolve().parents[2]
            / "system-design/kb-upgrade-v2/module-6-qfd/c1v_QFD.xlsx"
        )
        template = str(default_template)

    src = Path(template)
    if not src.exists():
        raise FileNotFoundError(
            f"QFD template not found at {src}. Pass options.templatePath to override."
        )

    out_path = output_dir / (options.get("outputFilename") or "qfd.xlsx")
    shutil.copyfile(src, out_path)

    wb = load_workbook(out_path)
    sheet_name = options.get("sheetName") or "QFD Template"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    writes = _plan_writes(instance)
    for cell, val in writes:
        ws[cell] = val
    wb.save(out_path)
    warnings.append(f"gen-qfd: wrote {len(writes)} cells into {out_path.name}")

    return [{"target": "xlsx", "path": str(out_path)}]


if __name__ == "__main__":
    raise SystemExit(run_generator(generator_name="gen-qfd", render_fn=render))
