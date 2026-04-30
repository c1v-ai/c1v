#!/usr/bin/env python3
"""gen-decision-net.py — Crawley Decision Network generator.

Renders ``decision_network.v1.json`` to:
  - xlsx  : scoring matrix (alternatives × criteria) via openpyxl
  - svg   : DAG graph of decisions/alternatives/outcomes via networkx+matplotlib
  - svg   : Pareto scatter (cost vs utility) via matplotlib
  - svg   : sensitivity heatmap (weight × alternative → score delta) via matplotlib
  - svg   : utility-vector bar chart per alternative via matplotlib

Schema: expects canonical ``decision-network.schema.json`` once T4b lands. Until
then falls back to the permissive stub at
``scripts/artifact-generators/common/schemas/decision-network.stub.schema.json``
(callers should set schemaRef to ``decision-network.stub.schema.json``).

Instance shape (loosely; permissive during Wave-1):
{
  "decisions":     [{"id": "D1", "question": "...", "dependsOn": ["D0"]}],
  "alternatives":  [{"id": "A1", "decisionId": "D1", "name": "..."}],
  "criteria":      [{"id": "C1", "name": "cost", "weight": 0.3, "direction": "min|max"}],
  "scores":        {"A1": {"C1": 0.8, "C2": 0.6}},
  "costs":         {"A1": 1200.0},
  "utilities":     {"A1": 0.74},
  "sensitivity":   {"A1": {"C1": -0.04, "C2": 0.11}},
  "winner":        "A1"
}

All sections are optional — missing sections produce warnings and skip their
corresponding artifact.
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from common import extender_init  # noqa: F401,E402  (mutates schema roots)
from common.runner import run_generator  # noqa: E402

# Matplotlib MUST be configured before import of pyplot
import matplotlib  # noqa: E402
matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
import networkx as nx  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _default_font_kwargs() -> dict[str, Any]:
    """Viewer-agnostic font settings: use matplotlib's built-in DejaVu Sans."""
    return {"family": "DejaVu Sans"}


def _save_svg(fig, path: Path) -> None:
    fig.savefig(path, format="svg", bbox_inches="tight")
    plt.close(fig)


def _xlsx_scoring_matrix(
    instance: dict, out_path: Path, warnings: list[str]
) -> bool:
    alts = instance.get("alternatives") or []
    criteria = instance.get("criteria") or []
    scores = instance.get("scores") or {}
    if not alts or not criteria:
        warnings.append(
            "gen-decision-net: xlsx skipped — missing alternatives or criteria"
        )
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Scoring Matrix"

    hdr_font = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="DDDDDD")

    ws.cell(row=1, column=1, value="Alternative").font = hdr_font
    ws.cell(row=1, column=1).fill = hdr_fill
    for j, c in enumerate(criteria, start=2):
        name = f"{c.get('name', c.get('id', f'C{j-1}'))} (w={c.get('weight', '?')})"
        cell = ws.cell(row=1, column=j, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
    ws.cell(row=1, column=len(criteria) + 2, value="Utility").font = hdr_font
    ws.cell(row=1, column=len(criteria) + 3, value="Cost").font = hdr_font

    utilities = instance.get("utilities") or {}
    costs = instance.get("costs") or {}
    winner = instance.get("winner")

    for i, a in enumerate(alts, start=2):
        aid = a.get("id", f"A{i-1}")
        label = a.get("name", aid)
        if winner == aid:
            label = f"★ {label}"
        ws.cell(row=i, column=1, value=label)
        alt_scores = scores.get(aid, {})
        for j, c in enumerate(criteria, start=2):
            cid = c.get("id")
            ws.cell(row=i, column=j, value=alt_scores.get(cid))
        ws.cell(row=i, column=len(criteria) + 2, value=utilities.get(aid))
        ws.cell(row=i, column=len(criteria) + 3, value=costs.get(aid))

    for col_idx in range(1, len(criteria) + 4):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

    wb.save(out_path)
    return True


def _svg_dag(instance: dict, out_path: Path, warnings: list[str]) -> bool:
    decisions = instance.get("decisions") or []
    alts = instance.get("alternatives") or []
    if not decisions and not alts:
        warnings.append("gen-decision-net: DAG svg skipped — no decisions or alternatives")
        return False

    g = nx.DiGraph()
    for d in decisions:
        did = d.get("id")
        g.add_node(did, kind="decision", label=d.get("question", did))
        for dep in d.get("dependsOn") or []:
            g.add_edge(dep, did)
    for a in alts:
        aid = a.get("id")
        g.add_node(aid, kind="alternative", label=a.get("name", aid))
        did = a.get("decisionId")
        if did:
            g.add_edge(did, aid)

    winner = instance.get("winner")
    # Simple layered layout: decisions on top rows, alternatives on bottom.
    try:
        pos = nx.nx_agraph.graphviz_layout(g, prog="dot")
    except Exception:
        pos = nx.spring_layout(g, seed=42)

    fig, ax = plt.subplots(figsize=(10, 6))
    decision_nodes = [n for n, data in g.nodes(data=True) if data.get("kind") == "decision"]
    alt_nodes = [n for n, data in g.nodes(data=True) if data.get("kind") == "alternative"]
    winner_nodes = [n for n in alt_nodes if n == winner]
    normal_alt = [n for n in alt_nodes if n != winner]

    nx.draw_networkx_edges(g, pos, ax=ax, arrows=True, edge_color="#888")
    nx.draw_networkx_nodes(
        g, pos, nodelist=decision_nodes, node_shape="s",
        node_color="#5998C5", node_size=1800, ax=ax,
    )
    nx.draw_networkx_nodes(
        g, pos, nodelist=normal_alt, node_shape="o",
        node_color="#FBFCFC", edgecolors="#0B2C29", node_size=1600, ax=ax,
    )
    if winner_nodes:
        nx.draw_networkx_nodes(
            g, pos, nodelist=winner_nodes, node_shape="o",
            node_color="#F18F01", edgecolors="#0B2C29",
            linewidths=2.5, node_size=1800, ax=ax,
        )
    labels = {n: (data.get("label") or n)[:24] for n, data in g.nodes(data=True)}
    nx.draw_networkx_labels(
        g, pos, labels=labels, font_size=7, ax=ax, font_family="DejaVu Sans"
    )
    ax.set_title("Decision Network (DAG)", **_default_font_kwargs())
    ax.set_axis_off()
    _save_svg(fig, out_path)
    return True


def _svg_pareto(instance: dict, out_path: Path, warnings: list[str]) -> bool:
    alts = instance.get("alternatives") or []
    utilities = instance.get("utilities") or {}
    costs = instance.get("costs") or {}
    pts = [(a, costs.get(a["id"]), utilities.get(a["id"])) for a in alts
           if a.get("id") in costs and a.get("id") in utilities]
    if not pts:
        warnings.append("gen-decision-net: pareto svg skipped — missing costs or utilities")
        return False
    fig, ax = plt.subplots(figsize=(8, 6))
    xs = [p[1] for p in pts]
    ys = [p[2] for p in pts]
    labels = [p[0].get("name", p[0].get("id", "?")) for p in pts]
    winner = instance.get("winner")
    colors = [
        "#F18F01" if p[0].get("id") == winner else "#5998C5" for p in pts
    ]
    ax.scatter(xs, ys, c=colors, s=120, edgecolors="#0B2C29", zorder=3)
    for x, y, lab in zip(xs, ys, labels):
        ax.annotate(lab, (x, y), xytext=(5, 5), textcoords="offset points",
                    fontsize=8, **_default_font_kwargs())
    ax.set_xlabel("Cost", **_default_font_kwargs())
    ax.set_ylabel("Utility", **_default_font_kwargs())
    ax.set_title("Pareto: Cost vs Utility", **_default_font_kwargs())
    ax.grid(True, linestyle="--", alpha=0.4)
    _save_svg(fig, out_path)
    return True


def _svg_sensitivity(instance: dict, out_path: Path, warnings: list[str]) -> bool:
    sensitivity = instance.get("sensitivity") or {}
    criteria = instance.get("criteria") or []
    alts = instance.get("alternatives") or []
    if not sensitivity or not criteria or not alts:
        warnings.append(
            "gen-decision-net: sensitivity svg skipped — missing sensitivity/criteria/alternatives"
        )
        return False
    alt_ids = [a.get("id") for a in alts]
    crit_ids = [c.get("id") for c in criteria]
    matrix = []
    for aid in alt_ids:
        row = [float((sensitivity.get(aid) or {}).get(cid, 0.0)) for cid in crit_ids]
        matrix.append(row)

    fig, ax = plt.subplots(figsize=(max(6, 0.7 * len(crit_ids)),
                                    max(4, 0.5 * len(alt_ids))))
    im = ax.imshow(matrix, cmap="RdBu_r", aspect="auto",
                   vmin=-max(0.01, max(abs(v) for row in matrix for v in row) or 0.01),
                   vmax=max(0.01, max(abs(v) for row in matrix for v in row) or 0.01))
    ax.set_xticks(range(len(crit_ids)))
    ax.set_xticklabels([c.get("name", c.get("id")) for c in criteria],
                       rotation=45, ha="right", **_default_font_kwargs())
    ax.set_yticks(range(len(alt_ids)))
    ax.set_yticklabels([a.get("name", a.get("id")) for a in alts],
                       **_default_font_kwargs())
    for i, row in enumerate(matrix):
        for j, v in enumerate(row):
            ax.text(j, i, f"{v:+.2f}", ha="center", va="center",
                    fontsize=7, color="#0B2C29", **_default_font_kwargs())
    ax.set_title("Sensitivity (Δscore per unit-weight shift)",
                 **_default_font_kwargs())
    fig.colorbar(im, ax=ax, fraction=0.02, pad=0.04)
    _save_svg(fig, out_path)
    return True


def _svg_utility_bars(instance: dict, out_path: Path, warnings: list[str]) -> bool:
    alts = instance.get("alternatives") or []
    utilities = instance.get("utilities") or {}
    if not alts or not utilities:
        warnings.append("gen-decision-net: utility bars svg skipped — missing utilities")
        return False
    labels = [a.get("name", a.get("id")) for a in alts]
    values = [float(utilities.get(a.get("id"), 0.0)) for a in alts]
    winner = instance.get("winner")
    colors = ["#F18F01" if a.get("id") == winner else "#5998C5" for a in alts]
    fig, ax = plt.subplots(figsize=(max(6, 0.8 * len(alts)), 5))
    ax.bar(labels, values, color=colors, edgecolor="#0B2C29")
    ax.set_ylabel("Utility", **_default_font_kwargs())
    ax.set_title("Utility per Alternative", **_default_font_kwargs())
    ax.set_ylim(0, max(1.0, max(values) * 1.1 if values else 1.0))
    ax.grid(True, axis="y", linestyle="--", alpha=0.4)
    for i, v in enumerate(values):
        ax.text(i, v + 0.01, f"{v:.2f}", ha="center", fontsize=8,
                **_default_font_kwargs())
    plt.setp(ax.get_xticklabels(), rotation=20, ha="right")
    _save_svg(fig, out_path)
    return True


# ---------------------------------------------------------------------------
# render
# ---------------------------------------------------------------------------

def render(instance, output_dir: Path, targets, options, warnings):
    base = options.get("outputBasename") or "decision_network"
    # Multiple svgs — treat 'svg' target as "emit all svg variants".
    generated: list[dict[str, str]] = []

    if "xlsx" in targets:
        out = output_dir / f"{base}.xlsx"
        if _xlsx_scoring_matrix(instance, out, warnings):
            generated.append({"target": "xlsx", "path": str(out)})

    if "svg" in targets:
        mapping = [
            ("dag", _svg_dag),
            ("pareto", _svg_pareto),
            ("sensitivity", _svg_sensitivity),
            ("utility-bars", _svg_utility_bars),
        ]
        for suffix, fn in mapping:
            out = output_dir / f"{base}.{suffix}.svg"
            try:
                if fn(instance, out, warnings):
                    generated.append({"target": "svg", "path": str(out)})
            except Exception as exc:
                warnings.append(
                    f"gen-decision-net: {suffix} svg failed: {type(exc).__name__}: {exc}"
                )

    if not generated:
        warnings.append("gen-decision-net: no outputs produced (check targets/instance)")
    return generated


if __name__ == "__main__":
    raise SystemExit(run_generator(generator_name="gen-decision-net", render_fn=render))
