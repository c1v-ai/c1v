#!/usr/bin/env python3
"""gen-self-app-flat-tables.py — flatten c1v self-application JSON into
xlsx workbooks for M1 + M2.

Reads:
  system-design/kb-upgrade-v2/module-1-defining-scope/data_flows.v1.json
  system-design/kb-upgrade-v2/module-2-requirements/requirements_table.json
  system-design/kb-upgrade-v2/module-2-requirements/constants.v2.json
  system-design/kb-upgrade-v2/module-2-requirements/nfrs.v2.json

Writes (under --output-dir):
  M1-data-flows.xlsx       — 1 sheet (DataFlows, 15 entries × N cols)
  M2-requirements.xlsx     — 3 sheets:
    Requirements (99 × N) | Constants (28 × N) | NFRs (26 × N)

The output is consumed by assemble-master-xlsx.py — see
master-workbook.manifest.json. The script is intentionally generic about
record fields: it inspects the first record to build column headers, so
upstream JSON additions surface automatically.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _bold_header(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True)
    cell.fill = PatternFill('solid', fgColor='DDDDDD')
    cell.alignment = Alignment(wrap_text=True, vertical='top')


def _set_widths(ws, widths: List[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _flatten_value(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (list, tuple)):
        if not v:
            return ''
        if all(isinstance(x, (str, int, float, bool)) for x in v):
            return ', '.join(str(x) for x in v)
        return json.dumps(v, ensure_ascii=False)
    if isinstance(v, dict):
        return json.dumps(v, ensure_ascii=False)
    return v


def write_sheet(wb: Workbook, sheet_name: str, records: List[Dict[str, Any]],
                preferred_order: List[str] | None = None) -> None:
    ws = wb.create_sheet(sheet_name)
    if not records:
        ws.cell(row=1, column=1, value=f'(no rows in {sheet_name})')
        return

    # Build column ordering: preferred_order first, then any remaining keys
    # in insertion order from the first record.
    seen: List[str] = []
    if preferred_order:
        for k in preferred_order:
            if any(k in r for r in records) and k not in seen:
                seen.append(k)
    for r in records:
        for k in r.keys():
            if k not in seen:
                seen.append(k)

    for i, h in enumerate(seen, start=1):
        _bold_header(ws.cell(row=1, column=i), h)
    for r_idx, rec in enumerate(records, start=2):
        for c_idx, key in enumerate(seen, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_flatten_value(rec.get(key)))
            ws.cell(row=r_idx, column=c_idx).alignment = Alignment(
                wrap_text=True, vertical='top'
            )
    # Reasonable default widths.
    widths = []
    for k in seen:
        if 'description' in k or 'requirement' in k or 'text' in k or 'rationale' in k or 'notes' in k:
            widths.append(60)
        elif 'name' in k or 'derived_from' in k or 'verification' in k:
            widths.append(30)
        else:
            widths.append(18)
    _set_widths(ws, widths)
    ws.freeze_panes = 'A2'


def write_m1(out_dir: Path) -> Path:
    src = Path('system-design/kb-upgrade-v2/module-1-defining-scope/data_flows.v1.json')
    d = json.loads(src.read_text(encoding='utf-8'))
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, 'DataFlows', d.get('entries') or [],
                preferred_order=['id', 'name', 'description', 'source', 'sink',
                                 'payload_shape', 'criticality', 'pii_class'])
    out_path = out_dir / 'M1-data-flows.xlsx'
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def write_m3(out_dir: Path) -> Path:
    src = Path('system-design/kb-upgrade-v2/module-3-ffbd/ffbd.v1.json')
    d = json.loads(src.read_text(encoding='utf-8'))
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, 'Functions', d.get('functions') or [],
                preferred_order=['id', 'name', 'description'])
    write_sheet(wb, 'Arrows', d.get('arrows') or [])
    write_sheet(wb, 'LogicGates', d.get('logic_gates') or [])
    write_sheet(wb, 'DataBlocks', d.get('data_blocks') or [])
    write_sheet(wb, 'CrossCutting', d.get('cross_cutting_pervasive') or [])
    out_path = out_dir / 'M3-ffbd.xlsx'
    wb.save(str(out_path))
    return out_path


def _flatten_decision_nodes(phases: Dict[str, Any]) -> List[Dict[str, Any]]:
    nodes = []
    for phase_key, phase in (phases or {}).items():
        if not isinstance(phase, dict):
            continue
        for n in phase.get('decision_nodes') or []:
            nodes.append({
                'phase': phase_key,
                'id': n.get('id'),
                'title': n.get('title'),
                'question': n.get('question'),
                'alternative_count': len(n.get('alternatives') or []),
                'criteria_count': len(n.get('criteria') or []),
            })
    return nodes


def _flatten_decision_scores(phases: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows = []
    for phase_key, phase in (phases or {}).items():
        if not isinstance(phase, dict):
            continue
        for n in phase.get('decision_nodes') or []:
            alt_names = {a.get('id'): a.get('name') for a in n.get('alternatives') or []}
            for s in n.get('scores') or []:
                rows.append({
                    'decision_id': n.get('id'),
                    'decision_title': n.get('title'),
                    'alternative_id': s.get('alternative_id'),
                    'alternative_name': alt_names.get(s.get('alternative_id')),
                    'criterion_id': s.get('criterion_id'),
                    'raw_value': s.get('raw_value'),
                    'normalized_value': s.get('normalized_value'),
                    'kb_source': (s.get('empirical_priors') or {}).get('ref'),
                    'sample_size': (s.get('empirical_priors') or {}).get('sample_size'),
                })
    return rows


def write_m4(out_dir: Path) -> Path:
    src = Path('system-design/kb-upgrade-v2/module-4-decision-matrix/decision_network.v1.json')
    d = json.loads(src.read_text(encoding='utf-8'))
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet('Summary')
    summary['A1'] = 'M4 — Decision Network'
    summary['A1'].font = Font(bold=True, size=16)
    summary['A3'] = 'Selected architecture'
    summary['B3'] = d.get('selected_architecture_id')
    summary['A4'] = 'Phases'
    summary['B4'] = ', '.join(sorted((d.get('phases') or {}).keys()))
    summary['A5'] = 'System'
    summary['B5'] = d.get('system_name')
    summary['A6'] = 'Decision-audit entries'
    summary['B6'] = len(d.get('decision_audit') or [])
    _set_widths(summary, [30, 80])

    nodes = _flatten_decision_nodes(d.get('phases') or {})
    write_sheet(wb, 'DecisionNodes', nodes,
                preferred_order=['phase', 'id', 'title', 'question'])
    scores = _flatten_decision_scores(d.get('phases') or {})
    write_sheet(wb, 'Scores', scores,
                preferred_order=['decision_id', 'alternative_id',
                                 'alternative_name', 'criterion_id',
                                 'raw_value', 'normalized_value', 'kb_source'])
    write_sheet(wb, 'Audit', d.get('decision_audit') or [],
                preferred_order=['ts', 'actor', 'action', 'phase', 'rationale'])

    out_path = out_dir / 'M4-decision-network.xlsx'
    wb.save(str(out_path))
    return out_path


def write_m5(out_dir: Path) -> Path:
    src = Path('system-design/kb-upgrade-v2/module-5-formfunction/form_function_map.v1.json')
    d = json.loads(src.read_text(encoding='utf-8'))
    wb = Workbook()
    wb.remove(wb.active)

    sections = [
        ('Forms', 'phase_1_form_inventory', 'forms'),
        ('Functions', 'phase_2_function_inventory', 'functions'),
        ('ConceptMatrix', 'phase_3_concept_mapping_matrix', 'mappings'),
        ('Scoring', 'phase_4_concept_quality_scoring', 'scores'),
        ('OperandProcess', 'phase_5_operand_process_catalog', 'entries'),
        ('Concepts', 'phase_6_concept_alternatives', 'concepts'),
    ]
    for sheet_name, phase_key, list_key in sections:
        phase = d.get(phase_key) or {}
        # Try multiple common field names for the list payload.
        records = []
        if isinstance(phase, dict):
            for k in (list_key, 'rows', 'items', 'entries'):
                if isinstance(phase.get(k), list) and phase.get(k):
                    records = phase[k]
                    break
            # If no list found but phase is shallow, embed as single row.
            if not records and any(not isinstance(v, (dict, list)) for v in phase.values()):
                records = [{k: v for k, v in phase.items() if not k.startswith('_')}]
        write_sheet(wb, sheet_name, records)

    out_path = out_dir / 'M5-form-function.xlsx'
    wb.save(str(out_path))
    return out_path


def write_m2(out_dir: Path) -> Path:
    reqs = json.loads(
        Path('system-design/kb-upgrade-v2/module-2-requirements/requirements_table.json')
        .read_text(encoding='utf-8')
    )
    consts = json.loads(
        Path('system-design/kb-upgrade-v2/module-2-requirements/constants.v2.json')
        .read_text(encoding='utf-8')
    )
    nfrs = json.loads(
        Path('system-design/kb-upgrade-v2/module-2-requirements/nfrs.v2.json')
        .read_text(encoding='utf-8')
    )

    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(
        wb, 'Requirements', reqs.get('requirements_table') or [],
        preferred_order=['index', 'requirement', 'abstract_function_name',
                         'source_ucbd', 'source_step', 'referenced_constants',
                         'cross_cutting_ref'],
    )
    write_sheet(
        wb, 'Constants', consts.get('constants') or [],
        preferred_order=['constant_name', 'value', 'unit', 'category',
                         'derived_from', 'source', 'owner', 'status', 'notes'],
    )
    write_sheet(
        wb, 'NFRs', nfrs.get('nfrs') or [],
        preferred_order=['req_id', 'text', 'requirement_class', 'derived_from',
                         'verification_method', 'rationale', 'status'],
    )

    out_path = out_dir / 'M2-requirements.xlsx'
    wb.save(str(out_path))
    return out_path


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument('--output-dir', required=True)
    args = ap.parse_args()
    out_dir = Path(args.output_dir).resolve()

    m1 = write_m1(out_dir)
    m2 = write_m2(out_dir)
    m3 = write_m3(out_dir)
    m4 = write_m4(out_dir)
    m5 = write_m5(out_dir)
    print(json.dumps({
        'm1': str(m1), 'm2': str(m2), 'm3': str(m3),
        'm4': str(m4), 'm5': str(m5),
    }, indent=2))


if __name__ == '__main__':
    main()
